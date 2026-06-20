"""Tests de la **persistencia incremental del corpus** (Fase A MEJORADO, ADR-0011).

Anclan lo esencial del MVP:

- Cada predicción exitosa guarda **1** fila en ``submissions`` y **N** en ``observations``
  (el corpus que crece), por **JSON** y por **Excel**, en **línea** y por **lote**.
- El ``client_id`` viene del header ``X-Client-Id`` (default ``"default"``) y NO toca el
  cuerpo del contrato.
- **Best-effort:** si la persistencia falla, la predicción se responde igual (200).

El repositorio se inyecta en memoria (``:memory:``), aislado del disco real.
"""

from __future__ import annotations

import time
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from spc.api.main import crear_app
from spc.service.repositorio import RepositorioPredicciones

COLS_HISTORY = ["date", "store_id", "product_id", "units_sold", "on_promotion", "transactions"]
_UMBRAL_FORZAR_LOTE = "1"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Fixtures: repositorio en memoria + app con ese repo inyectado
# ---------------------------------------------------------------------------
@pytest.fixture
def repo() -> RepositorioPredicciones:
    """Repositorio SQLite en memoria (efímero), inspeccionable en el test."""
    r = RepositorioPredicciones.crear(":memory:")
    yield r
    r.cerrar()


@pytest.fixture
def client_persistente(registro, repo) -> object:
    """``TestClient`` con el registro y el repositorio inyectados (persistencia activa)."""
    app = crear_app(registro=registro, repositorio=repo, cors_origins=["http://localhost:5173"])
    with TestClient(app) as c:
        yield c


# ---------------------------------------------------------------------------
# Helpers de lectura del corpus + construcción de Excel
# ---------------------------------------------------------------------------
def _consultar(repo: RepositorioPredicciones, sql: str, params: tuple = ()) -> list[dict]:
    cur = repo.conexion().execute(sql, params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, fila, strict=True)) for fila in cur.fetchall()]


def _submissions(repo: RepositorioPredicciones) -> list[dict]:
    return _consultar(repo, "SELECT * FROM submissions ORDER BY id")


def _xlsx_sales(historico: list[dict]) -> bytes:
    """``.xlsx`` válido de SALES (hojas history + parameters), igual que el canal real."""
    wb = Workbook()
    wb.remove(wb.active)
    ws_h = wb.create_sheet(title="history")
    ws_h.append(list(COLS_HISTORY))
    for h in historico:
        ws_h.append([h.get(c) for c in COLS_HISTORY])
    ws_p = wb.create_sheet(title="parameters")
    ws_p.append(["granularity", "horizon"])
    ws_p.append(["day", 5])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# JSON en línea
# ---------------------------------------------------------------------------
def test_json_online_guarda_submission_y_observations(client_persistente, repo, historico_contrato):
    """SALES por JSON (en línea): 1 submission + N observations con los valores correctos."""
    r = client_persistente.post(
        "/sales", json={"granularity": "day", "horizon": 5, "history": historico_contrato}
    )
    assert r.status_code == 200, r.text

    subs = _submissions(repo)
    assert len(subs) == 1
    sub = subs[0]
    assert sub["domain"] == "sales"
    assert sub["channel"] == "json"
    assert sub["mode"] == "online"
    assert sub["client_id"] == "default"
    assert sub["n_rows"] == len(historico_contrato)
    assert sub["model_version"]  # SALES siempre expone 'model' (de la metadata o el stem)

    assert repo.contar_observaciones() == len(historico_contrato)
    obs = _consultar(repo, "SELECT * FROM observations LIMIT 1")[0]
    assert obs["store_id"] in {"1", "2"}
    assert obs["product_id"] == "BEVERAGES"
    assert obs["submission_id"] == sub["id"]


def test_client_id_desde_header(client_persistente, repo, historico_contrato):
    """El header ``X-Client-Id`` etiqueta el corpus; sin header cae a ``default``."""
    client_persistente.post(
        "/sales",
        json={"granularity": "day", "horizon": 3, "history": historico_contrato},
        headers={"X-Client-Id": "acme"},
    )
    subs = _submissions(repo)
    assert subs[-1]["client_id"] == "acme"
    assert repo.contar_observaciones(client_id="acme") == len(historico_contrato)


# ---------------------------------------------------------------------------
# Canal Excel
# ---------------------------------------------------------------------------
def test_excel_se_persiste_con_canal_excel(client_persistente, repo, historico_contrato):
    """SALES por Excel: se guarda con ``channel='excel'`` (mismo flujo, distinta puerta)."""
    contenido = _xlsx_sales(historico_contrato)
    r = client_persistente.post(
        "/sales/excel",
        files={"file": ("sales.xlsx", contenido, XLSX)},
    )
    assert r.status_code == 200, r.text
    sub = _submissions(repo)[-1]
    assert sub["channel"] == "excel"
    assert sub["domain"] == "sales"
    assert sub["mode"] == "online"


# ---------------------------------------------------------------------------
# Modo por lote
# ---------------------------------------------------------------------------
def test_lote_se_persiste_con_mode_batch(client_persistente, repo, historico_contrato, monkeypatch):
    """Por encima del umbral, el envío se procesa por lote y se guarda con ``mode='batch'``."""
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    r = client_persistente.post(
        "/sales", json={"granularity": "day", "horizon": 3, "history": historico_contrato}
    )
    assert r.status_code == 202, r.text
    acuse = r.json()

    # Esperar a que el trabajo termine (la persistencia ocurre dentro del worker).
    deadline = time.time() + 60.0
    estado = None
    while time.time() < deadline:
        estado = client_persistente.get(acuse["status_url"]).json()
        if estado["status"] in ("done", "error"):
            break
        time.sleep(0.1)
    assert estado is not None and estado["status"] == "done", f"el trabajo no terminó: {estado}"

    sub = _submissions(repo)[-1]
    assert sub["mode"] == "batch"
    assert sub["domain"] == "sales"
    assert repo.contar_observaciones() == len(historico_contrato)


# ---------------------------------------------------------------------------
# Best-effort: un fallo de persistencia NO rompe la predicción
# ---------------------------------------------------------------------------
class _RepoQueFalla:
    """Repositorio falso cuyo ``registrar`` siempre falla (para probar el best-effort)."""

    def registrar(self, **_kwargs) -> int:
        raise RuntimeError("fallo simulado de la base de corpus")

    def cerrar(self) -> None:
        pass


def test_fallo_de_persistencia_no_rompe_la_prediccion(registro, historico_contrato):
    """Si la BD falla, la predicción sigue devolviendo 200 con su respuesta correcta."""
    app = crear_app(registro=registro, repositorio=_RepoQueFalla())
    with TestClient(app) as c:
        r = c.post(
            "/sales", json={"granularity": "day", "horizon": 3, "history": historico_contrato}
        )
    assert r.status_code == 200, r.text
    cuerpo = r.json()
    assert cuerpo["field"] == "sales" and "forecast" in cuerpo
