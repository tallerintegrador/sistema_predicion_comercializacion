"""Tests del **entrenamiento por cliente bajo demanda** (ADR-0013).

Anclan, con fixtures sintéticos (sin GPU ni datos reales), las garantías clave:

- Disparar el entrenamiento (Excel, mismo contrato) → ver el job → obtener la **comparación
  honesta** (candidato vs congelado vs baseline) y el veredicto de adopción.
- El **default (congelado) NO cambia** para quien no pulsa el botón.
- Un cliente con **poca historia** recibe un aviso honesto (no se entrena a ciegas).
- El modelo por cliente **solo se sirve si se adoptó** (superó al congelado); si no mejoró,
  se sigue con el congelado y se reporta (aquí se fuerza con el umbral de mejora).
- El **switch** de serving activa/desactiva el modelo por cliente.

El entrenamiento corre en el executor in-process de la app (separado del de lote); los
tests **pollean** el resultado como haría el frontend.
"""

from __future__ import annotations

import time
from io import BytesIO

import numpy as np
import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from spc.api.main import crear_app

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
COLS = ["date", "store_id", "product_id", "units_sold", "on_promotion", "transactions"]


# ---------------------------------------------------------------------------
# Helpers de datos y de subida
# ---------------------------------------------------------------------------
def _historia(n_dias: int, *, nivel: float, seed: int = 13) -> list[dict]:
    """Histórico (forma de contrato) de 2 puntos de venta, AR(1) + promoción.

    ``nivel`` fija la escala de las ventas: una escala **grande** (≈ ``nivel``) hace que el
    modelo congelado —entrenado a escala pequeña en los fixtures— subprediga y el candidato
    por cliente lo supere; es un escenario honesto de mejora medible.
    """
    rng = np.random.default_rng(seed)
    fechas = [d.date().isoformat() for d in pd.date_range("2017-01-01", periods=n_dias, freq="D")]
    filas: list[dict] = []
    for pv in ("1", "2"):
        prev = rng.uniform(nivel * 0.8, nivel * 1.2)
        for f in fechas:
            promo = int(rng.integers(0, 6))
            val = max(0.0, 0.8 * prev + 0.02 * nivel * promo + rng.normal(0, nivel * 0.04))
            prev = val
            filas.append(
                {
                    "date": f,
                    "store_id": pv,
                    "product_id": "BEVERAGES",
                    "units_sold": round(val, 1),
                    "on_promotion": promo,
                    "transactions": round(val * 1.5, 0),
                }
            )
    return filas


def _xlsx_sales(history: list[dict], horizon: int = 7) -> bytes:
    """Construye un ``.xlsx`` de SALES (hojas history + parameters) en memoria."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("history")
    ws.append(COLS)
    for h in history:
        ws.append([h.get(c) for c in COLS])
    wp = wb.create_sheet("parameters")
    wp.append(["granularity", "horizon"])
    wp.append(["day", horizon])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _entrenar(client, history, client_id, source: str | None = None) -> str:
    """Sube el Excel al endpoint de entrenamiento y devuelve el job_id (202)."""
    ruta = "/training/sales/excel" + (f"?source={source}" if source else "")
    r = client.post(
        ruta,
        files={"file": ("datos.xlsx", _xlsx_sales(history), XLSX)},
        headers={"X-Client-Id": client_id},
    )
    assert r.status_code == 202, r.text
    acuse = r.json()
    # El worker corre en paralelo: el acuse puede mostrar cualquier estado válido.
    assert acuse["status"] in ("queued", "running", "done", "error")
    assert acuse["client_id"] == client_id
    return acuse["job_id"]


def _esperar(client, job_id: str, timeout: float = 120.0) -> tuple[int, dict]:
    """Pollea ``/training/jobs/{id}/result`` hasta que termine (como el frontend)."""
    fin = time.time() + timeout
    while time.time() < fin:
        r = client.get(f"/training/jobs/{job_id}/result")
        if r.status_code != 202:
            return r.status_code, r.json()
        time.sleep(0.4)
    raise AssertionError("el entrenamiento no terminó dentro del timeout")


@pytest.fixture
def client_train(registro, tmp_path) -> TestClient:
    """`TestClient` con carpeta de modelos por cliente temporal (ADR-0013 habilitado)."""
    app = crear_app(
        registro=registro,
        client_models_dir=tmp_path / "clientes",
        cors_origins=["http://localhost:5173"],
    )
    with TestClient(app) as c:
        yield c


# ---------------------------------------------------------------------------
# 1. Disparar → job → comparación honesta + adopción + serving + switch
# ---------------------------------------------------------------------------
def test_entrenar_adopta_sirve_y_switch(client_train):
    """Flujo completo: el candidato supera al congelado → adoptado → servido → switch."""
    history = _historia(180, nivel=1000.0)
    job_id = _entrenar(client_train, history, "acme")
    status, body = _esperar(client_train, job_id)

    # --- Comparación honesta presente y veredicto = adoptado ---
    assert status == 200, body
    assert body["outcome"] == "adopted", body
    assert body["metric"] == "WAPE_recursivo"
    assert body["candidate"]["WAPE"] < body["frozen"]["WAPE"]  # mejora real
    assert body["beats_frozen"] is True
    assert body["model_version"].startswith("regresion_cliente_")

    # --- Estado: se sirve el modelo por cliente ---
    est = client_train.get("/training/sales/status", headers={"X-Client-Id": "acme"}).json()
    assert est["has_client_model"] is True
    assert est["serving_client_model"] is True
    assert est["adopted_version"] == 1

    # --- Predicción: ESE cliente recibe su modelo (model = regresion_cliente_...) ---
    payload = {"granularity": "day", "horizon": 5, "history": history}
    r = client_train.post("/sales", json=payload, headers={"X-Client-Id": "acme"})
    assert r.status_code == 200, r.text
    assert r.json()["model"].startswith("regresion_cliente_")

    # --- Switch OFF → vuelve al congelado ---
    r_off = client_train.post(
        "/training/sales/serving", json={"enabled": False}, headers={"X-Client-Id": "acme"}
    )
    assert r_off.status_code == 200 and r_off.json()["serving_client_model"] is False
    r2 = client_train.post("/sales", json=payload, headers={"X-Client-Id": "acme"})
    assert r2.json()["model"] == "regresion_v3"  # congelado de los fixtures

    # --- Switch ON → vuelve a servir el del cliente ---
    r_on = client_train.post(
        "/training/sales/serving", json={"enabled": True}, headers={"X-Client-Id": "acme"}
    )
    assert r_on.json()["serving_client_model"] is True
    r3 = client_train.post("/sales", json=payload, headers={"X-Client-Id": "acme"})
    assert r3.json()["model"].startswith("regresion_cliente_")


# ---------------------------------------------------------------------------
# 2. Default congelado intacto para quien no entrena
# ---------------------------------------------------------------------------
def test_default_congelado_intacto(client_train):
    """Un cliente que NO entrena recibe el modelo congelado, idéntico al camino por defecto."""
    payload = {"granularity": "day", "horizon": 5, "history": _historia(120, nivel=1000.0)}
    r_sin_header = client_train.post("/sales", json=payload)
    r_otro_cliente = client_train.post("/sales", json=payload, headers={"X-Client-Id": "nadie"})
    assert r_sin_header.status_code == 200 and r_otro_cliente.status_code == 200
    # Mismo modelo congelado y misma respuesta byte a byte (el client_id no altera el default).
    assert r_otro_cliente.json()["model"] == "regresion_v3"
    assert r_otro_cliente.json() == r_sin_header.json()

    est = client_train.get("/training/sales/status", headers={"X-Client-Id": "nadie"}).json()
    assert est["has_client_model"] is False and est["serving_client_model"] is False


# ---------------------------------------------------------------------------
# 3. Poca historia → aviso honesto (no se entrena a ciegas)
# ---------------------------------------------------------------------------
def test_poca_historia_aviso_honesto(client_train):
    """Con historia insuficiente, el experimento no entrena y lo reporta; sigue el congelado."""
    history = _historia(20, nivel=1000.0)  # 20 días < mínimo (60)
    job_id = _entrenar(client_train, history, "pyme_chica")
    status, body = _esperar(client_train, job_id)
    assert status == 200, body
    assert body["outcome"] == "insufficient_data"
    assert body["missing"], "debe listar qué requisitos faltan"

    # No hay modelo adoptado: la predicción sigue con el congelado.
    est = client_train.get("/training/sales/status", headers={"X-Client-Id": "pyme_chica"}).json()
    assert est["serving_client_model"] is False
    r = client_train.post(
        "/sales",
        json={"granularity": "day", "horizon": 3, "history": history},
        headers={"X-Client-Id": "pyme_chica"},
    )
    assert r.json()["model"] == "regresion_v3"


# ---------------------------------------------------------------------------
# 4. No mejora → no se adopta, se reporta, sigue el congelado
# ---------------------------------------------------------------------------
def test_no_mejora_no_se_adopta(client_train, monkeypatch):
    """Con un umbral de mejora inalcanzable, el candidato no se adopta y se reporta honestamente."""
    # Exige una mejora absurda (999 puntos de WAPE) → ninguna mejora real la alcanza.
    monkeypatch.setenv("SPC_CLIENT_ADJ_MIN_IMPROVEMENT", "999")
    history = _historia(180, nivel=1000.0)
    job_id = _entrenar(client_train, history, "exigente")
    status, body = _esperar(client_train, job_id)
    assert status == 200, body
    assert body["outcome"] == "not_adopted", body
    assert body["beats_frozen"] is False

    # No se sirve el modelo por cliente: sigue el congelado.
    est = client_train.get("/training/sales/status", headers={"X-Client-Id": "exigente"}).json()
    assert est["serving_client_model"] is False
    assert est["has_client_model"] is True  # se entrenó y guardó (auditoría), pero no se adoptó
    r = client_train.post(
        "/sales",
        json={"granularity": "day", "horizon": 3, "history": history},
        headers={"X-Client-Id": "exigente"},
    )
    assert r.json()["model"] == "regresion_v3"


# ---------------------------------------------------------------------------
# 5. Bordes: job inexistente y feature deshabilitada
# ---------------------------------------------------------------------------
def test_job_inexistente_404(client_train):
    """Un job_id desconocido devuelve 404 con el cuerpo de error uniforme."""
    r = client_train.get("/training/jobs/noexiste/result")
    assert r.status_code == 404
    assert r.json()["error"]["type"] == "not_found"


def test_entrenamiento_deshabilitado_503(registro, tmp_path, monkeypatch):
    """Con el ajuste por cliente desactivado, el endpoint de entrenamiento responde 503."""
    monkeypatch.setenv("SPC_CLIENT_ADJ_ENABLED", "0")
    app = crear_app(registro=registro, client_models_dir=tmp_path / "clientes")
    with TestClient(app) as c:
        r = c.post(
            "/training/sales/excel",
            files={"file": ("d.xlsx", _xlsx_sales(_historia(120, nivel=1000.0)), XLSX)},
            headers={"X-Client-Id": "x"},
        )
        assert r.status_code == 503, r.text
        # Y el serving sigue siendo el congelado (default intacto).
        r2 = c.post(
            "/sales",
            json={"granularity": "day", "horizon": 3, "history": _historia(120, nivel=1000.0)},
            headers={"X-Client-Id": "x"},
        )
        assert r2.json()["model"] == "regresion_v3"
