"""Tests de la API **agnóstica auto-entrenada** (ADR-0023): ``/auto/{sales,inventory,purchases}``.

Verifican que el sistema entrena el ganador al vuelo sobre un esquema **declarado por el
cliente** (columnas arbitrarias, otro rubro), pronostica/clasifica, reusa el modelo
cacheado cuando la data no cambia y degrada con errores claros (400) ante esquema/datos
inválidos.
"""

from __future__ import annotations

import io
import json

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_datos(rows: list[dict], items: list[dict] | None = None) -> bytes:
    """Arma un .xlsx con la hoja 'datos' (y opcional 'items') desde filas de dict."""
    wb = Workbook()
    ws = wb.active
    ws.title = "datos"
    cols = list(rows[0])
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    if items:
        wi = wb.create_sheet("items")
        icols = list(items[0])
        wi.append(icols)
        for it in items:
            wi.append([it.get(c) for c in icols])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _filas_agnosticas(series=("L1", "L2"), dias: int = 100) -> list[dict]:
    """Datos de OTRO rubro (columnas libres): ``fecha``, ``sucursal``, ``linea``,
    ``ventas``, ``descuento`` (conocido a futuro), ``trafico`` (solo pasado), ``region``."""
    rng = np.random.default_rng(11)
    filas: list[dict] = []
    for serie in series:
        base = 130 if serie == series[0] else 60
        for i in range(dias):
            d = (pd.Timestamp("2024-01-01") + pd.Timedelta(days=i)).date().isoformat()
            ventas = max(0.0, base * (1 + 0.3 * np.sin(2 * np.pi * i / 7)) + rng.normal(0, 6))
            filas.append({
                "fecha": d, "sucursal": serie, "linea": "X",
                "ventas": round(ventas, 1), "descuento": int(rng.integers(0, 4)),
                "trafico": round(ventas * 1.4, 0), "region": "norte" if serie == series[0] else "sur",
            })
    return filas


_SCHEMA = {
    "target": "ventas", "date": "fecha", "series_keys": ["sucursal", "linea"],
    "features": [
        {"name": "descuento", "type": "numeric", "known_future": True},
        {"name": "trafico", "type": "numeric", "known_future": False},
        {"name": "region", "type": "categorical", "known_future": True},
    ],
}


def test_auto_sales_entrena_y_pronostica(client) -> None:
    r = client.post("/auto/sales", json={"schema": _SCHEMA, "horizon": 7, "granularity": "day",
                                         "rows": _filas_agnosticas()})
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["field"] == "sales"
    assert j["training"]["winner_algorithm"]  # se eligió un ganador real
    assert j["training"]["trained_rows"] > 0
    # 2 series × 7 días.
    assert len(j["forecast"]) == 14
    fila = j["forecast"][0]
    assert {"sucursal", "linea", "date", "forecast_demand"} <= set(fila)
    assert fila["forecast_demand"] >= 0


def test_auto_sales_agrega_semanal(client) -> None:
    r = client.post("/auto/sales", json={"schema": _SCHEMA, "horizon": 14, "granularity": "week",
                                         "rows": _filas_agnosticas()})
    assert r.status_code == 200, r.text
    # week agrega: menos filas que el equivalente diario (14 días × 2 series).
    assert 0 < len(r.json()["forecast"]) < 28


def test_auto_sales_reusa_cache(client) -> None:
    payload = {"schema": _SCHEMA, "horizon": 5, "rows": _filas_agnosticas()}
    primera = client.post("/auto/sales", json=payload).json()
    assert primera["training"]["reused_cached_model"] is False
    segunda = client.post("/auto/sales", json=payload).json()
    assert segunda["training"]["reused_cached_model"] is True
    assert segunda["training"]["schema_signature"] == primera["training"]["schema_signature"]


def test_auto_inventory_clasifica_y_recomienda(client) -> None:
    items = [
        {"sucursal": "L1", "linea": "X", "current_stock": 40, "lead_time_days": 3},
        {"sucursal": "L2", "linea": "X", "current_stock": 500, "lead_time_days": 5},
    ]
    r = client.post("/auto/inventory", json={"schema": _SCHEMA, "rows": _filas_agnosticas(), "items": items})
    assert r.status_code == 200, r.text
    alerts = r.json()["alerts"]
    assert len(alerts) == 2
    for a in alerts:
        assert a["demand_class"] in ("high", "low")
        assert 0.0 <= a["high_demand_probability"] <= 1.0
        assert isinstance(a["stockout_risk"], bool)
        assert a["recommended_stock"] >= 0


def test_auto_purchases_repone(client) -> None:
    items = [{"sucursal": "L1", "linea": "X", "current_stock": 40,
              "lead_time_days": 3, "target_coverage_days": 7}]
    r = client.post("/auto/purchases", json={"schema": _SCHEMA, "rows": _filas_agnosticas(), "items": items})
    assert r.status_code == 200, r.text
    rec = r.json()["recommendation"]
    assert len(rec) == 1
    assert rec[0]["replenishment_quantity"] >= 0
    assert rec[0]["reorder_point"] >= 0


def test_auto_sales_esquema_invalido_400(client) -> None:
    bad = {**_SCHEMA, "target": "no_existe"}
    r = client.post("/auto/sales", json={"schema": bad, "horizon": 3, "rows": _filas_agnosticas()})
    assert r.status_code == 400
    assert "no_existe" in r.json()["error"]["message"]


def test_auto_sales_sin_fecha_400(client) -> None:
    bad = {"target": "ventas", "series_keys": ["sucursal"], "features": []}
    r = client.post("/auto/sales", json={"schema": bad, "horizon": 3, "rows": _filas_agnosticas()})
    assert r.status_code == 400
    assert "date" in r.json()["error"]["message"]


def test_auto_inventory_serie_sin_historico_400(client) -> None:
    items = [{"sucursal": "FANTASMA", "linea": "X", "current_stock": 10}]
    r = client.post("/auto/inventory", json={"schema": _SCHEMA, "rows": _filas_agnosticas(), "items": items})
    assert r.status_code == 400


@pytest.mark.parametrize("falta", ["lead_time_days", "target_coverage_days"])
def test_auto_purchases_item_incompleto_400(client, falta) -> None:
    item = {"sucursal": "L1", "linea": "X", "current_stock": 40,
            "lead_time_days": 3, "target_coverage_days": 7}
    item.pop(falta)
    r = client.post("/auto/purchases", json={"schema": _SCHEMA, "rows": _filas_agnosticas(), "items": [item]})
    assert r.status_code == 400


# --- Canal Excel ---
def test_auto_sales_template_descarga(client) -> None:
    r = client.post("/auto/sales/template", json={"schema": _SCHEMA})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX_MEDIA
    wb = load_workbook(io.BytesIO(r.content))
    assert "datos" in wb.sheetnames and "instrucciones" in wb.sheetnames
    cabecera = [c.value for c in next(wb["datos"].iter_rows())]
    # Las columnas de la plantilla son exactamente las del esquema declarado.
    assert cabecera == ["fecha", "sucursal", "linea", "ventas", "descuento", "trafico", "region"]


def test_auto_inventory_template_incluye_items(client) -> None:
    r = client.post("/auto/inventory/template", json={"schema": _SCHEMA})
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert "items" in wb.sheetnames
    # La hoja items viene pre-llenada con una fila por serie.
    assert wb["items"].max_row >= 2


def test_auto_template_viene_pre_llenada_sin_celdas_vacias(client) -> None:
    """La plantilla trae un dataset COMPLETO listo para subir: muchas filas, cero vacíos."""
    r = client.post("/auto/sales/template", json={"schema": _SCHEMA})
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["datos"]
    assert ws.max_row > 50  # varias series × muchos días (no 1-2 placeholders)
    vacias = sum(1 for row in ws.iter_rows(min_row=2) for c in row if c.value is None)
    assert vacias == 0


def test_auto_template_descargada_se_sube_tal_cual(client) -> None:
    """El Excel descargado se puede subir SIN editar y produce un pronóstico."""
    tpl = client.post("/auto/sales/template", json={"schema": _SCHEMA})
    r = client.post(
        "/auto/sales/excel",
        files={"file": ("plantilla.xlsx", tpl.content, _XLSX_MEDIA)},
        data={"schema": json.dumps(_SCHEMA), "horizon": "7"},
    )
    assert r.status_code == 200, r.text
    assert len(r.json()["forecast"]) > 0


def test_auto_sales_excel_entrena_y_pronostica(client) -> None:
    xlsx = _xlsx_datos(_filas_agnosticas())
    r = client.post(
        "/auto/sales/excel",
        files={"file": ("datos.xlsx", xlsx, _XLSX_MEDIA)},
        data={"schema": json.dumps(_SCHEMA), "horizon": "7", "granularity": "day"},
    )
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["training"]["winner_algorithm"]
    assert len(j["forecast"]) == 14


def test_auto_inventory_excel_usa_hoja_items(client) -> None:
    items = [
        {"sucursal": "L1", "linea": "X", "current_stock": 40, "lead_time_days": 3},
        {"sucursal": "L2", "linea": "X", "current_stock": 500, "lead_time_days": 5},
    ]
    xlsx = _xlsx_datos(_filas_agnosticas(), items=items)
    r = client.post(
        "/auto/inventory/excel",
        files={"file": ("datos.xlsx", xlsx, _XLSX_MEDIA)},
        data={"schema": json.dumps(_SCHEMA)},
    )
    assert r.status_code == 200, r.text
    assert len(r.json()["alerts"]) == 2


def test_auto_sales_excel_sin_filas_400(client) -> None:
    wb = Workbook()
    wb.active.title = "datos"
    wb.active.append(["fecha", "sucursal", "linea", "ventas", "descuento", "trafico", "region"])
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post(
        "/auto/sales/excel",
        files={"file": ("vacio.xlsx", buf.getvalue(), _XLSX_MEDIA)},
        data={"schema": json.dumps(_SCHEMA), "horizon": "7"},
    )
    assert r.status_code == 400
