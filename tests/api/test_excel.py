"""Tests del **canal Excel** (Fase 3.3): Excel es solo otra puerta al mismo contrato.

Anclan lo esencial:

- **Equivalencia JSON↔Excel**: los mismos datos por una y otra puerta producen el
  **mismo resultado** (la prueba ancla de "Excel es solo otra puerta").
- **Plantilla**: se descarga, es un ``.xlsx`` válido y trae las hojas esperadas con los
  encabezados del contrato (en inglés).
- **Errores claros y estructurados**: texto en columna numérica, columna obligatoria
  faltante, celda obligatoria vacía, hoja faltante y archivo demasiado grande devuelven
  el **mismo cuerpo de error** que el JSON, indicando hoja/fila/columna.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

# Tipo MIME de un .xlsx (el endpoint lo exige al subir).
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Columnas del bloque history (orden del contrato).
COLS_HISTORY = ["date", "store_id", "product_id", "units_sold", "on_promotion", "transactions"]


def _xlsx(hojas: dict[str, list[list]]) -> bytes:
    """Construye un ``.xlsx`` en memoria: ``{nombre_hoja: [fila0(cabecera), fila1, ...]}``."""
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, filas in hojas.items():
        ws = wb.create_sheet(title=nombre)
        for fila in filas:
            ws.append(fila)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _history_filas(historico: list[dict]) -> list[list]:
    """Cabecera + filas del bloque history a partir del histórico del contrato."""
    filas = [list(COLS_HISTORY)]
    for h in historico:
        filas.append([h.get(c) for c in COLS_HISTORY])
    return filas


def _subir(client, ruta: str, contenido: bytes, data: dict | None = None):
    """Sube un .xlsx al endpoint indicado (multipart), con campos de formulario opcionales."""
    return client.post(ruta, files={"file": ("datos.xlsx", contenido, XLSX)}, data=data or {})


# Configuración del pronóstico de SALES: viaja como campos de formulario (no en el
# archivo), porque la plantilla de Ventas es solo-datos (ADR-0022).
def _subir_sales(client, contenido: bytes, granularity: str = "day", horizon: int = 5):
    """Sube el .xlsx de SALES con la configuración de pantalla (granularity/horizon)."""
    return _subir(client, "/sales/excel", contenido, {"granularity": granularity, "horizon": horizon})


def _assert_422_excel(r, hoja_fragmento: str) -> None:
    """422 controlado (type=validation) con un detalle que cita la hoja indicada."""
    assert r.status_code == 422, r.text
    cuerpo = r.json()
    assert cuerpo["error"]["type"] == "validation"
    campos = " ".join(d["field"] for d in cuerpo["error"]["details"])
    assert hoja_fragmento in campos, f"esperaba '{hoja_fragmento}' en {campos}"


# ---------------------------------------------------------------------------
# Plantilla: descarga y forma
# ---------------------------------------------------------------------------
@pytest.mark.parametrize(
    "dominio,hojas_datos",
    [
        ("sales", {"history"}),  # Ventas es solo-datos: sin hoja de parámetros (ADR-0022).
        ("purchases", {"history", "replenishment_params"}),
        ("inventory", {"history", "inventory_status"}),
    ],
)
def test_plantilla_se_descarga_y_es_valida(client, dominio, hojas_datos):
    """La plantilla se descarga como .xlsx válido, con las hojas y encabezados esperados."""
    r = client.get(f"/{dominio}/template")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == XLSX
    assert "attachment" in r.headers["content-disposition"]

    wb = load_workbook(BytesIO(r.content))
    assert hojas_datos <= set(wb.sheetnames)
    assert "instructions" in wb.sheetnames  # hoja de ayuda en español
    # Encabezados del bloque history en inglés (nombres canónicos del contrato).
    cabecera = [c.value for c in wb["history"][1]]
    assert cabecera[:4] == ["date", "store_id", "product_id", "units_sold"]


def test_plantilla_sales_es_solo_datos(client):
    """La plantilla de Ventas NO trae hoja de parámetros: la configuración va en pantalla."""
    r = client.get("/sales/template")
    wb = load_workbook(BytesIO(r.content))
    assert "parameters" not in wb.sheetnames
    assert set(wb.sheetnames) == {"instructions", "history"}


def test_plantilla_dominio_inexistente_es_404(client):
    """Un dominio no soportado no expone endpoint de plantilla."""
    assert client.get("/unknown/template").status_code == 404


# ---------------------------------------------------------------------------
# Equivalencia JSON ↔ Excel (la prueba ancla)
# ---------------------------------------------------------------------------
def test_equivalencia_sales(client, historico_contrato):
    """Mismos datos por JSON y por Excel → MISMA respuesta en SALES.

    La configuración (granularity/horizon) llega como campos de formulario (pantalla),
    no en el archivo: la plantilla de Ventas es solo-datos (ADR-0022).
    """
    payload = {"granularity": "day", "horizon": 5, "history": historico_contrato}
    r_json = client.post("/sales", json=payload)
    assert r_json.status_code == 200, r_json.text

    contenido = _xlsx({"history": _history_filas(historico_contrato)})
    r_xlsx = _subir_sales(client, contenido, granularity="day", horizon=5)
    assert r_xlsx.status_code == 200, r_xlsx.text
    assert r_xlsx.json() == r_json.json()


def test_sales_config_desde_la_peticion_manda(client, historico_contrato):
    """El horizonte enviado por formulario (pantalla) determina el resultado del Excel."""
    contenido = _xlsx({"history": _history_filas(historico_contrato)})
    r = _subir_sales(client, contenido, granularity="day", horizon=3)
    assert r.status_code == 200, r.text
    # 3 períodos diarios por serie: el horizonte de pantalla gobierna, no el archivo.
    fechas = {item["date"] for item in r.json()["forecast"]}
    assert len(fechas) == 3


def test_equivalencia_purchases(client, historico_contrato):
    """Mismos datos por JSON y por Excel → MISMA respuesta en PURCHASES."""
    params = [
        {
            "store_id": "1",
            "product_id": "BEVERAGES",
            "current_stock": 900,
            "lead_time_days": 3,
            "target_coverage_days": 7,
        }
    ]
    payload = {"history": historico_contrato, "replenishment_params": params}
    r_json = client.post("/purchases", json=payload)
    assert r_json.status_code == 200, r_json.text

    cols = ["store_id", "product_id", "current_stock", "lead_time_days", "target_coverage_days"]
    contenido = _xlsx(
        {
            "history": _history_filas(historico_contrato),
            "replenishment_params": [cols, [params[0][c] for c in cols]],
        }
    )
    r_xlsx = _subir(client, "/purchases/excel", contenido)
    assert r_xlsx.status_code == 200, r_xlsx.text
    assert r_xlsx.json() == r_json.json()


def test_equivalencia_inventory(client, historico_contrato):
    """Mismos datos por JSON y por Excel → MISMA respuesta en INVENTORY."""
    inv = [{"store_id": "1", "product_id": "BEVERAGES", "current_stock": 300, "lead_time_days": 3}]
    payload = {"history": historico_contrato, "inventory_status": inv}
    r_json = client.post("/inventory", json=payload)
    assert r_json.status_code == 200, r_json.text

    cols = ["store_id", "product_id", "current_stock", "lead_time_days"]
    contenido = _xlsx(
        {
            "history": _history_filas(historico_contrato),
            "inventory_status": [cols, [inv[0][c] for c in cols]],
        }
    )
    r_xlsx = _subir(client, "/inventory/excel", contenido)
    assert r_xlsx.status_code == 200, r_xlsx.text
    assert r_xlsx.json() == r_json.json()


def test_equivalencia_descarga_y_sube_la_propia_plantilla(client):
    """La plantilla descargada (solo datos) se sube con la config de pantalla y predice."""
    contenido = client.get("/sales/template").content
    r = _subir_sales(client, contenido)
    assert r.status_code == 200, r.text
    assert r.json()["field"] == "sales"


def test_fecha_como_celda_datetime_equivale_a_iso(client, historico_contrato):
    """Una fecha escrita como celda de fecha (datetime) se convierte a ISO igual que el JSON."""
    from datetime import date

    payload = {"granularity": "day", "horizon": 3, "history": historico_contrato}
    r_json = client.post("/sales", json=payload)

    filas = [list(COLS_HISTORY)]
    for h in historico_contrato:
        fila = [h.get(c) for c in COLS_HISTORY]
        fila[0] = date.fromisoformat(h["date"])  # celda de fecha real, no texto ISO
        filas.append(fila)
    contenido = _xlsx({"history": filas})
    r_xlsx = _subir_sales(client, contenido, horizon=3)
    assert r_xlsx.status_code == 200, r_xlsx.text
    assert r_xlsx.json() == r_json.json()


# ---------------------------------------------------------------------------
# Errores claros y estructurados (hoja/fila/columna)
# ---------------------------------------------------------------------------
def test_error_texto_en_columna_numerica(client, historico_contrato):
    """Texto no numérico en units_sold → 422 citando history y la fila."""
    filas = _history_filas(historico_contrato)
    filas[1][COLS_HISTORY.index("units_sold")] = "muchas"  # no convertible
    contenido = _xlsx({"history": filas})
    r = _subir_sales(client, contenido)
    _assert_422_excel(r, "history.row2.units_sold")


@pytest.mark.parametrize(
    "dominio,hoja,cols,fila",
    [
        (
            "purchases",
            "replenishment_params",
            ["store_id", "product_id", "current_stock", "lead_time_days", "target_coverage_days"],
            ["1", "BEVERAGES", 900, 3, 7],
        ),
        (
            "inventory",
            "inventory_status",
            ["store_id", "product_id", "current_stock", "lead_time_days"],
            ["1", "BEVERAGES", 300, 3],
        ),
    ],
)
def test_error_texto_en_columna_numerica_por_dominio(
    client, historico_contrato, dominio, hoja, cols, fila
):
    """Texto no numérico en la hoja propia de PURCHASES/INVENTORY → 422 citando hoja/fila/col."""
    mala = list(fila)
    mala[cols.index("current_stock")] = "mucho"  # no convertible a número
    contenido = _xlsx({"history": _history_filas(historico_contrato), hoja: [cols, mala]})
    r = _subir(client, f"/{dominio}/excel", contenido)
    _assert_422_excel(r, f"{hoja}.row2.current_stock")


def test_error_columna_obligatoria_faltante(client, historico_contrato):
    """Falta la columna obligatoria units_sold → 422 indicando la columna faltante."""
    cols = [c for c in COLS_HISTORY if c != "units_sold"]
    filas = [cols] + [[h.get(c) for c in cols] for h in historico_contrato]
    contenido = _xlsx({"history": filas})
    r = _subir_sales(client, contenido)
    _assert_422_excel(r, "history.row1.units_sold")


def test_error_celda_obligatoria_vacia(client, historico_contrato):
    """Celda vacía en un campo obligatorio (store_id) → 422 citando hoja/fila/columna."""
    filas = _history_filas(historico_contrato)
    filas[1][COLS_HISTORY.index("store_id")] = None
    contenido = _xlsx({"history": filas})
    r = _subir_sales(client, contenido)
    _assert_422_excel(r, "history.row2.store_id")


def test_error_hoja_faltante(client, historico_contrato):
    """Falta la hoja history (única de SALES) → 422 indicando la hoja faltante."""
    # Un .xlsx sin la hoja history: la configuración va por formulario, pero los datos no.
    contenido = _xlsx({"otra_cosa": [["a"], [1]]})
    r = _subir_sales(client, contenido)
    _assert_422_excel(r, "history")


def test_error_columna_desconocida(client, historico_contrato):
    """Una columna fuera del contrato → 422 (coherente con extra='forbid' del JSON)."""
    filas = _history_filas(historico_contrato)
    filas[0].append("sabotaje")  # encabezado desconocido
    for f in filas[1:]:
        f.append(1)
    contenido = _xlsx({"history": filas})
    r = _subir_sales(client, contenido)
    _assert_422_excel(r, "sabotaje")


def test_error_regla_de_contrato_horizon(client, historico_contrato):
    """horizon=0 (viola gt=0) llega por formulario y se valida con el MISMO modelo → 422."""
    contenido = _xlsx({"history": _history_filas(historico_contrato)})
    r = _subir_sales(client, contenido, horizon=0)
    _assert_422_excel(r, "horizon")


def test_error_archivo_ilegible(client):
    """Un archivo que no es .xlsx → 422 controlado (no un volcado de pila)."""
    r = _subir_sales(client, b"esto no es un excel")
    assert r.status_code == 422, r.text
    assert r.json()["error"]["type"] == "validation"


def test_error_archivo_demasiado_grande(client, monkeypatch, historico_contrato):
    """Un archivo por encima del tope configurable → 413 controlado."""
    import spc.api.routers.excel as mod

    monkeypatch.setattr(mod, "excel_max_bytes", lambda: 100)  # tope diminuto para la prueba
    contenido = _xlsx({"history": _history_filas(historico_contrato)})
    r = _subir_sales(client, contenido)
    assert r.status_code == 413, r.text
    assert r.json()["error"]["type"] == "invalid_request"
