"""Tests del **modo por lote** (Fase 3.4): ruteo por nº de filas, in-process.

Anclan lo esencial del modo de ejecución:

- **En línea (chico):** por debajo del umbral, respuesta síncrona inmediata (200), como hoy.
- **Por lote (grande):** por encima del umbral, acuse 202 con ``job_id``; el estado pasa
  a ``done`` y el resultado se recupera.
- **Equivalencia en línea↔lote:** el MISMO dato por ambos modos da el **mismo resultado**.
- **Umbral configurable:** ``SPC_ONLINE_MAX_ROWS`` decide la frontera (se respeta).
- **``job_id`` inexistente:** error claro 404 con el cuerpo uniforme.

El umbral se baja por variable de entorno para forzar el modo lote con el fixture
sintético pequeño, sin fabricar históricos gigantes (``online_max_rows()`` lee el
entorno en cada petición).
"""

from __future__ import annotations

import time
from io import BytesIO

import pytest
from openpyxl import Workbook

# Umbral artificialmente alto/bajo para fijar el modo sin construir datos enormes.
_UMBRAL_ALTO = "1000000"
_UMBRAL_FORZAR_LOTE = "1"

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
COLS_HISTORY = ["date", "store_id", "product_id", "units_sold", "on_promotion", "transactions"]


def _xlsx_sales(historico: list[dict]) -> bytes:
    """Construye un ``.xlsx`` válido de SALES (solo datos: hoja history) en memoria.

    La configuración (granularity/horizon) ya no viaja en el archivo: se envía como
    campos de formulario en la petición (ADR-0022).
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws_h = wb.create_sheet(title="history")
    ws_h.append(list(COLS_HISTORY))
    for h in historico:
        ws_h.append([h.get(c) for c in COLS_HISTORY])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Configuración del pronóstico de SALES como campos de formulario (pantalla).
_SALES_FORM = {"granularity": "day", "horizon": 5}


def _payloads(historico: list[dict]) -> dict[str, tuple[str, dict]]:
    """Petición válida por dominio (ruta + cuerpo)."""
    return {
        "sales": ("/sales", {"granularity": "day", "horizon": 5, "history": historico}),
        "purchases": (
            "/purchases",
            {
                "history": historico,
                "replenishment_params": [
                    {
                        "store_id": "1",
                        "product_id": "BEVERAGES",
                        "current_stock": 900,
                        "lead_time_days": 3,
                        "target_coverage_days": 7,
                    }
                ],
            },
        ),
        "inventory": (
            "/inventory",
            {
                "history": historico,
                "inventory_status": [
                    {"store_id": "1", "product_id": "BEVERAGES", "current_stock": 300, "lead_time_days": 3}
                ],
            },
        ),
    }


def _payloads_sin_historico(historico: list[dict]) -> dict[str, tuple[str, dict]]:
    """Petición **válida en esquema** pero con un producto sin histórico (→ regla de negocio 400).

    Sirve para comprobar que ese 400 se comporta igual en línea y por lote: el bloque
    ``history`` es legítimo (decide el ruteo por volumen), pero el producto consultado
    no aparece en él, así que la lógica de negocio no puede pronosticarlo.
    """
    return {
        "purchases": (
            "/purchases",
            {
                "history": historico,
                "replenishment_params": [
                    {
                        "store_id": "99",
                        "product_id": "NO_EXISTE",
                        "current_stock": 10,
                        "lead_time_days": 2,
                        "target_coverage_days": 3,
                    }
                ],
            },
        ),
        "inventory": (
            "/inventory",
            {
                "history": historico,
                "inventory_status": [
                    {"store_id": "1", "product_id": "NO_EXISTE", "current_stock": 10}
                ],
            },
        ),
    }


def _xlsx_dominio(historico: list[dict], hoja: str, cols: list[str], fila: list) -> bytes:
    """``.xlsx`` válido de un dominio: hoja ``history`` + la hoja propia (1 fila de datos)."""
    wb = Workbook()
    wb.remove(wb.active)
    ws_h = wb.create_sheet(title="history")
    ws_h.append(list(COLS_HISTORY))
    for h in historico:
        ws_h.append([h.get(c) for c in COLS_HISTORY])
    ws = wb.create_sheet(title=hoja)
    ws.append(list(cols))
    ws.append(list(fila))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _esperar_resultado(client, acuse: dict, timeout: float = 60.0) -> object:
    """Sondea el estado hasta ``done``/``error`` y devuelve la respuesta de /result."""
    deadline = time.time() + timeout
    estado = None
    while time.time() < deadline:
        estado = client.get(acuse["status_url"]).json()
        if estado["status"] in ("done", "error"):
            break
        time.sleep(0.1)
    assert estado is not None and estado["status"] == "done", f"el trabajo no terminó: {estado}"
    return client.get(acuse["result_url"])


def _esperar_estado_terminal(client, acuse: dict, timeout: float = 60.0) -> dict:
    """Sondea el estado hasta ``done``/``error`` (sin exigir éxito) y lo devuelve."""
    deadline = time.time() + timeout
    estado = None
    while time.time() < deadline:
        estado = client.get(acuse["status_url"]).json()
        if estado["status"] in ("done", "error"):
            break
        time.sleep(0.1)
    assert estado is not None and estado["status"] in ("done", "error"), f"el trabajo no terminó: {estado}"
    return estado


# ---------------------------------------------------------------------------
# En línea (chico): comportamiento síncrono de siempre, intacto
# ---------------------------------------------------------------------------
def test_envio_chico_responde_en_linea_200(client, historico_contrato):
    """Por debajo del umbral por defecto, SALES responde 200 síncrono (sin job_id)."""
    ruta, payload = _payloads(historico_contrato)["sales"]
    r = client.post(ruta, json=payload)
    assert r.status_code == 200, r.text
    cuerpo = r.json()
    assert cuerpo["field"] == "sales" and "forecast" in cuerpo
    assert "job_id" not in cuerpo


# ---------------------------------------------------------------------------
# Por lote (grande): 202 + job_id → estado → resultado
# ---------------------------------------------------------------------------
def test_envio_grande_devuelve_202_y_job_id(client, historico_contrato, monkeypatch):
    """Por encima del umbral, el mismo endpoint acusa 202 con job_id y enlaces."""
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    ruta, payload = _payloads(historico_contrato)["sales"]
    r = client.post(ruta, json=payload)
    assert r.status_code == 202, r.text
    acuse = r.json()
    assert acuse["mode"] == "batch"
    assert acuse["domain"] == "sales"
    assert acuse["rows"] == len(historico_contrato)
    assert acuse["job_id"] and acuse["status"] in ("queued", "running", "done")
    assert acuse["status_url"].endswith(acuse["job_id"])
    assert acuse["result_url"].endswith(f"{acuse['job_id']}/result")


def test_envio_grande_completa_y_entrega_resultado(client, historico_contrato, monkeypatch):
    """El trabajo pasa a 'done' y /result entrega 200 con la respuesta del dominio."""
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    ruta, payload = _payloads(historico_contrato)["sales"]
    acuse = client.post(ruta, json=payload).json()

    res = _esperar_resultado(client, acuse)
    assert res.status_code == 200, res.text
    cuerpo = res.json()
    assert cuerpo["field"] == "sales"
    assert cuerpo["forecast"], "la lista de pronóstico vino vacía"


def test_estado_reporta_metadatos_del_trabajo(client, historico_contrato, monkeypatch):
    """GET /jobs/{id} expone dominio, filas y marcas de tiempo coherentes."""
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    ruta, payload = _payloads(historico_contrato)["inventory"]
    acuse = client.post(ruta, json=payload).json()
    _esperar_resultado(client, acuse)

    estado = client.get(acuse["status_url"]).json()
    assert estado["domain"] == "inventory"
    assert estado["rows"] == len(historico_contrato)
    assert estado["status"] == "done"
    assert estado["created_at"] and estado["finished_at"]


# ---------------------------------------------------------------------------
# Equivalencia en línea ↔ lote: el mismo dato da el mismo resultado
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("dominio", ["sales", "purchases", "inventory"])
def test_equivalencia_en_linea_y_lote(client, historico_contrato, monkeypatch, dominio):
    """El MISMO dato por ambos modos produce el MISMO cuerpo de respuesta."""
    ruta, payload = _payloads(historico_contrato)[dominio]

    # En línea (umbral alto → síncrono).
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_ALTO)
    r_online = client.post(ruta, json=payload)
    assert r_online.status_code == 200, r_online.text

    # Por lote (umbral mínimo → asíncrono), mismo payload.
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    acuse = client.post(ruta, json=payload).json()
    r_lote = _esperar_resultado(client, acuse)
    assert r_lote.status_code == 200, r_lote.text

    assert r_lote.json() == r_online.json()


# ---------------------------------------------------------------------------
# Umbral configurable: se respeta la frontera
# ---------------------------------------------------------------------------
def test_umbral_configurable_se_respeta(client, historico_contrato, monkeypatch):
    """El mismo payload va en línea o a lote según SPC_ONLINE_MAX_ROWS."""
    ruta, payload = _payloads(historico_contrato)["sales"]

    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_ALTO)
    assert client.post(ruta, json=payload).status_code == 200

    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    assert client.post(ruta, json=payload).status_code == 202


def test_umbral_en_la_frontera_exacta_va_en_linea(client, historico_contrato, monkeypatch):
    """`len(history) == umbral` se procesa en línea (la frontera es inclusiva)."""
    ruta, payload = _payloads(historico_contrato)["sales"]
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", str(len(historico_contrato)))
    assert client.post(ruta, json=payload).status_code == 200

    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", str(len(historico_contrato) - 1))
    assert client.post(ruta, json=payload).status_code == 202


# ---------------------------------------------------------------------------
# job_id inexistente → 404 claro
# ---------------------------------------------------------------------------
def test_estado_de_job_inexistente_da_404(client):
    """GET /jobs/{id} con un id desconocido devuelve 404 con el cuerpo de error uniforme."""
    r = client.get("/jobs/no-existe-1234")
    assert r.status_code == 404, r.text
    cuerpo = r.json()
    assert cuerpo["error"]["type"] == "not_found"
    assert "no-existe-1234" in cuerpo["error"]["message"]


def test_resultado_de_job_inexistente_da_404(client):
    """GET /jobs/{id}/result con un id desconocido también devuelve 404 claro."""
    r = client.get("/jobs/no-existe-1234/result")
    assert r.status_code == 404, r.text
    assert r.json()["error"]["type"] == "not_found"


# ---------------------------------------------------------------------------
# Regla de negocio en lote → 400 idéntico al de en línea (mismo error por ambos modos)
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("dominio", ["purchases", "inventory"])
def test_regla_de_negocio_en_lote_da_el_mismo_400_que_en_linea(
    client, historico_contrato, monkeypatch, dominio
):
    """Un producto sin histórico falla igual por lote que en línea: 400 invalid_request.

    En línea devuelve **400** de inmediato. Por lote el envío se ACEPTA (202, porque el
    bloque history es grande y legítimo), el trabajo termina en estado ``error`` y
    ``GET /result`` reproduce el **mismo** 400 (tipo y mensaje), no un 500 ni un cuerpo
    distinto. Ancla que el worker mapea la regla de negocio igual que la capa en línea.
    """
    ruta, payload = _payloads_sin_historico(historico_contrato)[dominio]

    # En línea (umbral alto → síncrono): la regla de negocio da 400 al instante.
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_ALTO)
    r_online = client.post(ruta, json=payload)
    assert r_online.status_code == 400, r_online.text
    err_online = r_online.json()["error"]
    assert err_online["type"] == "invalid_request"

    # Por lote (umbral mínimo → asíncrono): se acepta (202), termina en 'error'…
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    acuse = client.post(ruta, json=payload).json()
    assert acuse["mode"] == "batch" and acuse["domain"] == dominio
    estado = _esperar_estado_terminal(client, acuse)
    assert estado["status"] == "error"

    # … y /result reproduce el MISMO 400 (mismo tipo y mensaje que en línea).
    res = client.get(acuse["result_url"])
    assert res.status_code == 400, res.text
    err_lote = res.json()["error"]
    assert err_lote["type"] == "invalid_request"
    assert err_lote["message"] == err_online["message"]


# ---------------------------------------------------------------------------
# El canal Excel también rutea por volumen (mismo umbral, misma frontera)
# ---------------------------------------------------------------------------
def test_excel_grande_tambien_rutea_a_lote(client, historico_contrato, monkeypatch):
    """`POST /sales/excel` por encima del umbral devuelve 202 (mismo ruteo que el JSON)."""
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    contenido = _xlsx_sales(historico_contrato)
    r = client.post(
        "/sales/excel", files={"file": ("datos.xlsx", contenido, XLSX)}, data=_SALES_FORM
    )
    assert r.status_code == 202, r.text
    acuse = r.json()
    assert acuse["mode"] == "batch" and acuse["domain"] == "sales"
    assert acuse["rows"] == len(historico_contrato)


def test_excel_chico_sigue_en_linea(client, historico_contrato, monkeypatch):
    """Por debajo del umbral, el Excel responde 200 síncrono (sin job_id), como hoy."""
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_ALTO)
    contenido = _xlsx_sales(historico_contrato)
    r = client.post(
        "/sales/excel", files={"file": ("datos.xlsx", contenido, XLSX)}, data=_SALES_FORM
    )
    assert r.status_code == 200, r.text
    assert r.json()["field"] == "sales"


@pytest.mark.parametrize(
    "dominio,hoja,cols,fila",
    [
        (
            "purchases",
            "replenishment_params",
            ["store_id", "product_id", "current_stock", "lead_time_days", "target_coverage_days"],
            ["1", "BEVERAGES", 900, 3, 7],
        ),
        (
            "inventory",
            "inventory_status",
            ["store_id", "product_id", "current_stock", "lead_time_days"],
            ["1", "BEVERAGES", 300, 3],
        ),
    ],
)
def test_excel_grande_rutea_a_lote_por_dominio(
    client, historico_contrato, monkeypatch, dominio, hoja, cols, fila
):
    """PURCHASES e INVENTORY por Excel rutean a lote igual que SALES (mismo umbral)."""
    monkeypatch.setenv("SPC_ONLINE_MAX_ROWS", _UMBRAL_FORZAR_LOTE)
    contenido = _xlsx_dominio(historico_contrato, hoja, cols, fila)
    r = client.post(f"/{dominio}/excel", files={"file": ("datos.xlsx", contenido, XLSX)})
    assert r.status_code == 202, r.text
    acuse = r.json()
    assert acuse["mode"] == "batch" and acuse["domain"] == dominio
    assert acuse["rows"] == len(historico_contrato)
