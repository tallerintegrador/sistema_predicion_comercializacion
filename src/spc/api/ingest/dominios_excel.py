"""Canal **Excel** del contrato 3×3 (``/v2``): plantilla/ejemplo + lectura.

A diferencia del canal agnóstico (columnas libres), aquí el formato es **fijo por
dominio** (``spc.synthetic.esquemas``). Este módulo:

- ``generar_excel`` — arma un ``.xlsx`` con la hoja **datos** (cabecera exacta + filas de
  ejemplo) y una hoja **instrucciones** (el diccionario de variables en lenguaje simple).
- ``leer_excel`` — lee un ``.xlsx`` subido de vuelta a ``rows`` (lista de dicts) en el
  orden del esquema, para alimentar el motor 3×3.

Openpyxl ya es dependencia de la API (canal Excel agnóstico).
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from spc.service import onboarding
from spc.service.errores import SolicitudInvalida
from spc.synthetic.esquemas import esquema_de

HOJA_DATOS = "datos"
HOJA_INSTRUCCIONES = "instrucciones"

_AZUL = "2F6F8F"
_FONT_CABECERA = Font(bold=True, color="FFFFFF")
_FILL_CABECERA = PatternFill("solid", fgColor=_AZUL)
_FONT_TITULO = Font(bold=True, size=13, color="1F4E5F")
_FONT_SUB = Font(bold=True, color="1F4E5F")
_WRAP = Alignment(vertical="top", wrap_text=True)


def _hoja_datos(wb: Workbook, dominio: str, filas: list[dict[str, Any]]) -> None:
    """Hoja 'datos': cabecera con las columnas del esquema + las filas de ejemplo."""
    esq = esquema_de(dominio)
    ws = wb.active
    ws.title = HOJA_DATOS
    for j, nombre in enumerate(esq.orden, start=1):
        celda = ws.cell(row=1, column=j, value=nombre)
        celda.font = _FONT_CABECERA
        celda.fill = _FILL_CABECERA
        ws.column_dimensions[get_column_letter(j)].width = max(12, len(nombre) + 2)
    for i, fila in enumerate(filas, start=2):
        for j, nombre in enumerate(esq.orden, start=1):
            ws.cell(row=i, column=j, value=fila.get(nombre))
    ws.freeze_panes = "A2"


def _hoja_instrucciones(wb: Workbook, dominio: str) -> None:
    """Hoja 'instrucciones': diccionario de variables + qué predice cada modelo.

    Por cada columna explica **qué es**, **para qué le sirve al sistema** (objetivo /
    factor / identificador / la calcula el sistema), si es **obligatoria u opcional** y —si
    se calcula sola— **su fórmula**. Todo se deriva del diccionario (fuente única).
    """
    dicc = onboarding.diccionario_de(dominio)
    ws = wb.create_sheet(HOJA_INSTRUCCIONES)
    ws.cell(row=1, column=1, value=f"Formato de {dominio.upper()} — {dicc['formato']}").font = _FONT_TITULO

    ws.cell(row=3, column=1, value="Qué te va a predecir el sistema:").font = _FONT_SUB
    qp = dicc["que_se_predice"]
    lineas = [
        f"• Pronóstico (regresión): {qp['regresion']['explicacion']}",
        f"• Alerta (clasificación): {qp['clasificacion']['explicacion']}",
        f"• Grupos (clustering): {qp['clustering']['explicacion']}",
    ]
    for k, texto in enumerate(lineas, start=4):
        ws.cell(row=k, column=1, value=texto)

    ws.cell(
        row=8, column=1,
        value=(
            "Nota: las columnas marcadas «Opcional» las calcula el sistema con su fórmula "
            "si no las traes; puedes dejarlas en blanco o no incluirlas."
        ),
    ).alignment = _WRAP

    encabezados = ["Columna", "Tipo", "¿Obligatoria?", "Qué es", "Para qué le sirve al sistema", "Ejemplo"]
    fila0 = 10
    ws.cell(row=fila0 - 1, column=1, value="Columnas del formato:").font = _FONT_SUB
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(row=fila0, column=j, value=h)
        c.font = _FONT_CABECERA
        c.fill = _FILL_CABECERA
    for i, col in enumerate(dicc["columnas"], start=fila0 + 1):
        para_que = col["uso"]
        if col.get("formula"):
            para_que = f"{para_que}  ·  Fórmula: {col['formula']}"
        ws.cell(row=i, column=1, value=col["nombre"])
        ws.cell(row=i, column=2, value=col["tipo"])
        ws.cell(row=i, column=3, value="Sí" if col["obligatoria"] else "Opcional")
        ws.cell(row=i, column=4, value=col["descripcion"]).alignment = _WRAP
        ws.cell(row=i, column=5, value=para_que).alignment = _WRAP
        ws.cell(row=i, column=6, value=str(col["ejemplo"]))
    anchos = [26, 16, 14, 46, 62, 16]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def generar_excel(dominio: str, filas: list[dict[str, Any]]) -> bytes:
    """Devuelve el ``.xlsx`` (bytes) con la hoja de datos + la de instrucciones."""
    esquema_de(dominio)  # valida el dominio (lanza si es desconocido)
    wb = Workbook()
    _hoja_datos(wb, dominio, filas)
    _hoja_instrucciones(wb, dominio)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def leer_excel(contenido: bytes, dominio: str) -> list[dict[str, Any]]:
    """Lee un ``.xlsx`` subido a ``rows`` (lista de dicts) en el orden del esquema."""
    esq = esquema_de(dominio)
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl lanza varios tipos ante un archivo corrupto
        raise SolicitudInvalida("No se pudo leer el archivo Excel (¿está dañado o no es .xlsx?).") from exc

    ws = wb[HOJA_DATOS] if HOJA_DATOS in wb.sheetnames else wb.worksheets[0]
    iterador = ws.iter_rows(values_only=True)
    try:
        cabecera = [str(c).strip() if c is not None else "" for c in next(iterador)]
    except StopIteration:
        raise SolicitudInvalida("El Excel está vacío: falta la cabecera con las columnas.") from None

    # Solo se exigen las columnas OBLIGATORIAS; las calculadas ausentes las rellena el motor
    # con su fórmula (spc.service.motor_3x3.construir_dataframe).
    faltan = [c for c in esq.columnas_obligatorias() if c not in cabecera]
    if faltan:
        raise SolicitudInvalida(
            f"Al Excel de '{dominio}' le faltan columnas obligatorias: {', '.join(faltan)}."
        )
    presentes = [c for c in esq.orden if c in cabecera]
    indice = {nombre: cabecera.index(nombre) for nombre in presentes}
    fechas = {c.nombre for c in esq.columnas if c.tipo == "date"}

    filas: list[dict[str, Any]] = []
    for cruda in iterador:
        if cruda is None or all(v is None for v in cruda):
            continue  # salta filas vacías
        fila: dict[str, Any] = {}
        for nombre in presentes:
            valor = cruda[indice[nombre]] if indice[nombre] < len(cruda) else None
            if nombre in fechas and isinstance(valor, (date, datetime)):
                valor = valor.date().isoformat() if isinstance(valor, datetime) else valor.isoformat()
            fila[nombre] = valor
        filas.append(fila)

    if not filas:
        raise SolicitudInvalida("El Excel no tiene filas de datos debajo de la cabecera.")
    return filas
