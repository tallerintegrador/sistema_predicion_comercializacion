"""Canal **Excel** para la predicción agnóstica (ADR-0023): plantilla + lectura.

A diferencia del Excel retail (`ingest/plantilla.py`), aquí las columnas **no son fijas**:
se derivan del `SchemaSpec` que el cliente declara. Se genera una plantilla con las
columnas exactas de su esquema (fecha, claves de serie, objetivo y features) + una hoja de
instrucciones en español, y se lee de vuelta el ``.xlsx`` a ``rows``/``items`` genéricos
(lista de dicts de columnas libres) para alimentar el mismo servicio agnóstico.
"""

from __future__ import annotations

import math
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from spc.api.schemas.agnostico import SchemaSpec
from spc.service.errores import SolicitudInvalida

HOJA_DATOS = "datos"
HOJA_ITEMS = "items"
HOJA_INSTRUCCIONES = "instrucciones"

# La plantilla viene **pre-llenada** con un dataset completo y listo para predecir
# (todas las columnas, varias series, historia suficiente). El usuario puede subirla tal
# cual para ver el sistema funcionando, o reemplazar las filas por sus datos reales.
_N_SERIES = 3
_DIAS = 90
_FECHA_INICIO = date(2024, 1, 1)
_CATEGORIAS = ("tipo_A", "tipo_B", "tipo_C")

_AZUL = "2F6F8F"
_GRIS = "EEF3F6"
_FONT_TITULO = Font(bold=True, size=13, color="1F4E5F")
_FONT_CABECERA = Font(bold=True, color="FFFFFF")
_FILL_CABECERA = PatternFill("solid", fgColor=_AZUL)
_FONT_SUB = Font(bold=True, color="1F4E5F")
_FILL_SUB = PatternFill("solid", fgColor=_GRIS)
_WRAP = Alignment(vertical="top", wrap_text=True)

# Columnas reservadas de la hoja `items` (inventario/compras), según el dominio.
_ITEMS_BASE = ["current_stock", "lead_time_days"]
_ITEMS_COMPRAS = ["current_stock", "lead_time_days", "target_coverage_days"]


def _columnas_datos(schema: SchemaSpec) -> list[tuple[str, str]]:
    """Lista ordenada ``(columna, tipo_humano)`` de la hoja de datos según el esquema."""
    cols: list[tuple[str, str]] = []
    if schema.date:
        cols.append((schema.date, "fecha (YYYY-MM-DD)"))
    for k in schema.series_keys:
        cols.append((k, "texto"))
    cols.append((schema.target, "número (lo que se predice)"))
    for f in schema.features:
        tipo = "número" if f.type == "numeric" else "texto"
        if f.type == "numeric" and f.known_future is False:
            tipo += " (solo histórico)"
        cols.append((f.name, tipo))
    return cols


def _valor_serie(schema: SchemaSpec, key: str, s: int) -> str:
    """Valor sintético de una clave de serie para la serie ``s`` (legible y distinto)."""
    return f"{key}_{s + 1}"


def _filas_completas(schema: SchemaSpec) -> list[dict[str, Any]]:
    """Dataset **completo y predecible** que llena TODAS las columnas del esquema.

    Genera ``_N_SERIES`` series × ``_DIAS`` días con estacionalidad semanal, fin de
    semana y un efecto de promoción/feriado, de modo que el archivo se pueda subir tal
    cual y el modelo entrene y prediga (no son celdas placeholder: son datos reales).
    """
    cats = [f for f in schema.features if f.type == "categorical"]
    num_kf = [f for f in schema.features if f.type == "numeric" and f.known_future]
    num_pp = [f for f in schema.features if f.type == "numeric" and not f.known_future]

    filas: list[dict[str, Any]] = []
    for s in range(_N_SERIES):
        base = 140.0 - 35.0 * s  # series con niveles distintos (140, 105, 70)
        for i in range(_DIAS):
            d = _FECHA_INICIO + timedelta(days=i)
            finde = 0.15 if d.weekday() >= 5 else 0.0
            # Ruido determinista (reproducible, sin depender de numpy).
            ruido = 6.0 * math.sin(i * 1.7 + s) * math.cos(i * 0.3)

            fila: dict[str, Any] = {}
            if schema.date:
                fila[schema.date] = d.isoformat()
            for k in schema.series_keys:
                fila[k] = _valor_serie(schema, k, s)

            # Categóricas: rotan; primera = "señal" que también empuja la demanda.
            cat_signal = 0.0
            for j, f in enumerate(cats):
                etiqueta = _CATEGORIAS[(i + j) % len(_CATEGORIAS)]
                fila[f.name] = etiqueta
                if j == 0 and etiqueta == _CATEGORIAS[0]:
                    cat_signal = 5.0

            # Numéricas conocidas a futuro: valores realistas según el nombre (heurística);
            # la promoción (si la hay) empuja la demanda.
            promo = 0.0
            es_feriado = 1.0 if d.day in (1, 15) else 0.0
            for j, f in enumerate(num_kf):
                nl = f.name.lower()
                if any(t in nl for t in ("feriado", "festivo", "holiday", "evento", "event")):
                    val: Any = int(es_feriado)
                elif any(t in nl for t in ("promo", "descuento", "oferta", "discount")):
                    val = i % 3  # 0/1/2
                    promo = float(val)
                elif any(t in nl for t in ("preci", "price", "costo", "cost", "tarifa")):
                    val = round(9.5 - 0.5 * (i % 3), 2)  # baja con la promo
                elif any(t in nl for t in ("temp", "clima", "weather", "grados")):
                    val = round(18.0 + 8.0 * math.sin(2 * math.pi * i / 365.0), 1)
                else:
                    val = i % 4  # numérica genérica pequeña
                    if j == 0 and promo == 0.0:
                        promo = float(val)
                fila[f.name] = val

            # Objetivo: estacionalidad semanal + finde + promo + señal categórica.
            objetivo = max(
                0.0,
                base * (1 + 0.25 * math.sin(2 * math.pi * i / 7.0) + finde)
                + 8.0 * promo
                + cat_signal
                + ruido,
            )
            fila[schema.target] = round(objetivo, 1)

            # Numéricas solo-pasado: correlacionan con el objetivo (p. ej. tráfico).
            for j, f in enumerate(num_pp):
                fila[f.name] = round(objetivo * (1.3 + 0.1 * j), 1)

            filas.append(fila)
    return filas


def _items_completos(schema: SchemaSpec, extra: list[str]) -> list[dict[str, Any]]:
    """Una fila por serie con stock y tiempos (hoja ``items`` lista para usar)."""
    filas: list[dict[str, Any]] = []
    for s in range(_N_SERIES):
        fila: dict[str, Any] = {k: _valor_serie(schema, k, s) for k in schema.series_keys}
        fila["current_stock"] = 200 - 40 * s
        fila["lead_time_days"] = 3
        if "target_coverage_days" in extra:
            fila["target_coverage_days"] = 7
        filas.append(fila)
    return filas


def _escribir_datos(ws: Any, cols: list[tuple[str, str]], ejemplos: list[dict[str, Any]]) -> None:
    nombres = [c[0] for c in cols]
    for j, nombre in enumerate(nombres, start=1):
        c = ws.cell(row=1, column=j, value=nombre)
        c.font = _FONT_CABECERA
        c.fill = _FILL_CABECERA
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(nombre) + 4)
    for i, ej in enumerate(ejemplos, start=2):
        for j, nombre in enumerate(nombres, start=1):
            ws.cell(row=i, column=j, value=ej.get(nombre))
    ws.freeze_panes = "A2"


def _escribir_items(ws: Any, schema: SchemaSpec, extra: list[str], filas: list[dict[str, Any]]) -> None:
    cols = [*schema.series_keys, *extra]
    for j, nombre in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=nombre)
        c.font = _FONT_CABECERA
        c.fill = _FILL_CABECERA
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(nombre) + 4)
    for i, fila in enumerate(filas, start=2):
        for j, nombre in enumerate(cols, start=1):
            ws.cell(row=i, column=j, value=fila.get(nombre))
    ws.freeze_panes = "A2"


def _escribir_instrucciones(ws: Any, schema: SchemaSpec, dominio: str) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 60
    fila = 1
    ws.cell(row=fila, column=1, value=f"Plantilla a tu medida — {dominio}").font = _FONT_TITULO
    fila += 2
    notas = [
        "Cómo usar esta plantilla:",
        "• Viene PRE-LLENADA con un ejemplo COMPLETO y listo para subir (puedes probarla tal cual).",
        "• Para tus datos: reemplaza las filas de la hoja 'datos' por las tuyas. NO cambies los nombres de las columnas.",
        "• Las columnas salen de TU esquema declarado en pantalla (objetivo, fecha, series y features).",
        "• Fechas en formato ISO YYYY-MM-DD (p. ej. 2024-01-01).",
        "• Cuantos más días e historia por serie, mejor predice (ideal: 40+ días por serie).",
    ]
    if dominio in ("inventory", "purchases"):
        notas.append("• Hoja 'items': una fila por serie con su stock y tiempos de entrega (también pre-llenada).")
    notas.append("• Sube el archivo en pantalla; el sistema entrena el mejor modelo y predice.")
    for n in notas:
        ws.cell(row=fila, column=1, value=n)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        fila += 1
    fila += 1

    ws.cell(row=fila, column=1, value="Columna").font = _FONT_SUB
    ws.cell(row=fila, column=2, value="Rol").font = _FONT_SUB
    ws.cell(row=fila, column=3, value="Detalle").font = _FONT_SUB
    for j in range(1, 4):
        ws.cell(row=fila, column=j).fill = _FILL_SUB
    fila += 1

    def _fila(col: str, rol: str, det: str) -> None:
        nonlocal fila
        ws.cell(row=fila, column=1, value=col)
        ws.cell(row=fila, column=2, value=rol)
        d = ws.cell(row=fila, column=3, value=det)
        d.alignment = _WRAP
        fila += 1

    if schema.date:
        _fila(schema.date, "fecha", "Índice temporal. Conocido a futuro (calendario).")
    for k in schema.series_keys:
        _fila(k, "clave de serie", "Identifica cada serie (p. ej. local, producto).")
    _fila(schema.target, "objetivo", "Lo que se predice. Su historia genera la señal base.")
    for f in schema.features:
        if f.type == "categorical":
            rol = "feature categórica"
            det = "Etiqueta/segmento."
        elif f.known_future is False:
            rol = "feature (solo histórico)"
            det = "No se conoce a futuro (tráfico, reservas): solo se usan sus rezagos."
        else:
            rol = "feature (conocida a futuro)"
            det = "Su valor del período a predecir se conoce (promo, precio, feriado)."
        _fila(f.name, rol, det)


def generar_plantilla(schema: SchemaSpec, dominio: str) -> bytes:
    """Genera el ``.xlsx`` del esquema declarado.

    La hoja ``datos`` va **primera y activa** para que, al abrir el archivo, se vean de
    inmediato **todas** las columnas del esquema (objetivo, fecha, claves de serie y cada
    feature). Luego van ``items`` (si aplica) y ``instrucciones``.
    """
    wb = Workbook()
    ws_datos = wb.active
    ws_datos.title = HOJA_DATOS
    cols = _columnas_datos(schema)
    _escribir_datos(ws_datos, cols, _filas_completas(schema))

    if dominio in ("inventory", "purchases"):
        extra = _ITEMS_COMPRAS if dominio == "purchases" else _ITEMS_BASE
        _escribir_items(wb.create_sheet(HOJA_ITEMS), schema, extra, _items_completos(schema, extra))

    _escribir_instrucciones(wb.create_sheet(HOJA_INSTRUCCIONES), schema, dominio)
    wb.active = 0  # abrir en la hoja de datos (muestra todas las columnas)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _normalizar(valor: Any) -> Any:
    """Normaliza un valor de celda: fechas a ISO, el resto tal cual."""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    return valor


def _leer_hoja(wb: Any, nombre: str) -> list[dict[str, Any]]:
    if nombre not in wb.sheetnames:
        return []
    ws = wb[nombre]
    filas = ws.iter_rows(values_only=True)
    try:
        cabecera = next(filas)
    except StopIteration:
        return []
    columnas = [str(c) for c in cabecera if c is not None]
    out: list[dict[str, Any]] = []
    for fila in filas:
        if fila is None or all(v is None for v in fila):
            continue
        registro = {col: _normalizar(val) for col, val in zip(columnas, fila, strict=False) if col}
        if any(v is not None for v in registro.values()):
            out.append(registro)
    return out


def leer_libro(contenido: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Lee el ``.xlsx`` y devuelve ``(rows, items)`` (items vacío si no hay esa hoja)."""
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - frontera de entrada: traducir a 400 claro
        raise SolicitudInvalida(f"No se pudo leer el archivo Excel: {exc}") from exc
    # La hoja de datos puede llamarse 'datos' o ser la primera no-instrucciones.
    rows = _leer_hoja(wb, HOJA_DATOS)
    if not rows:
        for nombre in wb.sheetnames:
            if nombre not in (HOJA_INSTRUCCIONES, HOJA_ITEMS):
                rows = _leer_hoja(wb, nombre)
                if rows:
                    break
    if not rows:
        raise SolicitudInvalida("El Excel no contiene filas de datos en la hoja 'datos'.")
    items = _leer_hoja(wb, HOJA_ITEMS)
    return rows, items


__all__ = ["generar_plantilla", "leer_libro", "HOJA_DATOS", "HOJA_ITEMS"]
