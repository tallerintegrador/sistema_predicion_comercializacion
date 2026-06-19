"""Lector de Excel: ``.xlsx`` subido → petición validada del contrato.

Es el corazón del canal: aplica el **punto técnico clave** de la Fase 3.1 — como la
validación es ``strict``, las celdas (que Excel suele entregar como texto o como
decimales tipo ``5.0``) se **convierten explícitamente** al tipo del contrato
**antes** de validar. Luego se arma exactamente la misma petición que enviaría el
JSON y se valida con los **mismos modelos Pydantic**.

Cualquier problema (archivo ilegible, hoja o columna faltante, celda vacía en campo
obligatorio, celda no convertible o regla del contrato incumplida) se reúne en un
``ErrorExcel`` que indica **hoja, fila y columna**, y que la capa API traduce al
**mismo cuerpo de error** (HTTP 422) que usa el JSON.
"""

from __future__ import annotations

from datetime import date as Date
from datetime import datetime as DateTime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ValidationError

from spc.api.ingest.esquema_excel import HojaExcel, PlantillaDominio, plantilla_de
from spc.api.schemas.almacen import AlmacenRequest
from spc.api.schemas.compras import ComprasRequest
from spc.api.schemas.comunes import DetalleError
from spc.api.schemas.ventas import VentasRequest

# Petición Pydantic por dominio (los MISMOS modelos strict que valida el JSON).
MODELO_PETICION: dict[str, type[BaseModel]] = {
    "sales": VentasRequest,
    "purchases": ComprasRequest,
    "inventory": AlmacenRequest,
}

_VERDADEROS = {"true", "1", "yes", "sí", "si", "verdadero", "y"}
_FALSOS = {"false", "0", "no", "falso", "n"}


class ErrorExcel(Exception):
    """Entrada por Excel mal formada: reúne los problemas con su hoja/fila/columna.

    La capa API lo traduce al cuerpo de error uniforme (HTTP 422), igual que un error
    de validación del JSON.
    """

    def __init__(self, detalles: list[DetalleError], mensaje: str | None = None) -> None:
        self.detalles = detalles
        self.mensaje = mensaje or "El archivo Excel no cumple el contrato de datos."
        super().__init__(self.mensaje)


class ArchivoDemasiadoGrande(Exception):
    """El ``.xlsx`` subido supera el tope de tamaño (``SPC_EXCEL_MAX_BYTES``).

    La capa API lo traduce a HTTP 413. Es un límite plano de protección; el ruteo por
    volumen (en línea/lote) es de la Fase 3.4 y aquí no se implementa.
    """


# ---------------------------------------------------------------------------
# Conversión explícita de una celda al tipo del contrato
# ---------------------------------------------------------------------------
def _es_vacia(valor: Any) -> bool:
    """True si la celda está vacía (``None`` o texto en blanco)."""
    return valor is None or (isinstance(valor, str) and valor.strip() == "")


class _CeldaInvalida(ValueError):
    """La celda no se puede convertir al tipo esperado (mensaje legible en español)."""


def convertir_celda(valor: Any, conversor: str) -> Any:
    """Convierte una celda al tipo del contrato. Lanza ``_CeldaInvalida`` si no puede.

    Implementa la conversión explícita exigida por ``strict``: produce el tipo Python
    **exacto** (entero real, ``float``, ``bool``, fecha ISO o id de texto) que los
    modelos aceptan sin coerción.
    """
    if conversor == "date":
        return _a_fecha_iso(valor)
    if conversor == "id":
        return _a_id(valor)
    if conversor == "number":
        return _a_numero(valor)
    if conversor == "integer":
        return _a_entero(valor)
    if conversor == "boolean":
        return _a_booleano(valor)
    # "enum": se pasa como texto; el modelo valida el conjunto de valores permitidos.
    return str(valor).strip()


def _a_fecha_iso(valor: Any) -> str:
    """Fecha → cadena ISO ``YYYY-MM-DD`` (Excel suele dar ``datetime``)."""
    if isinstance(valor, DateTime):
        return valor.date().isoformat()
    if isinstance(valor, Date):
        return valor.isoformat()
    if isinstance(valor, str):
        try:
            return Date.fromisoformat(valor.strip()).isoformat()
        except ValueError as exc:
            raise _CeldaInvalida("no es una fecha ISO (use YYYY-MM-DD).") from exc
    raise _CeldaInvalida("no es una fecha (use una celda de fecha o texto YYYY-MM-DD).")


def _a_id(valor: Any) -> str:
    """Identificador → texto (``5.0`` → ``\"5\"``; nunca ``\"5.0\"``)."""
    if isinstance(valor, bool):
        raise _CeldaInvalida("un identificador no puede ser booleano.")
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float):
        return str(int(valor)) if valor.is_integer() else str(valor)
    return str(valor).strip()


def _a_numero(valor: Any) -> float:
    """Número (decimal) a ``float``; acepta texto numérico (``\"900\"`` → ``900.0``)."""
    if isinstance(valor, bool):
        raise _CeldaInvalida("se esperaba un número, no un booleano.")
    if isinstance(valor, int | float):
        return float(valor)
    if isinstance(valor, str):
        try:
            return float(valor.strip())
        except ValueError as exc:
            raise _CeldaInvalida("no es un número.") from exc
    raise _CeldaInvalida("no es un número.")


def _a_entero(valor: Any) -> int:
    """Entero exacto; ``5.0`` → ``5`` pero ``5.5`` se rechaza; texto numérico admitido."""
    if isinstance(valor, bool):
        raise _CeldaInvalida("se esperaba un entero, no un booleano.")
    if isinstance(valor, int):
        return valor
    if isinstance(valor, float):
        if valor.is_integer():
            return int(valor)
        raise _CeldaInvalida("no es un entero (tiene parte decimal).")
    if isinstance(valor, str):
        try:
            f = float(valor.strip())
        except ValueError as exc:
            raise _CeldaInvalida("no es un entero.") from exc
        if f.is_integer():
            return int(f)
        raise _CeldaInvalida("no es un entero (tiene parte decimal).")
    raise _CeldaInvalida("no es un entero.")


def _a_booleano(valor: Any) -> bool:
    """Booleano; acepta ``TRUE/FALSE`` y texto común (``true``/``false``/``sí``/``no``)."""
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, int | float) and valor in (0, 1):
        return bool(valor)
    if isinstance(valor, str):
        t = valor.strip().lower()
        if t in _VERDADEROS:
            return True
        if t in _FALSOS:
            return False
    raise _CeldaInvalida("no es un booleano (use true/false).")


# ---------------------------------------------------------------------------
# Lectura de una hoja → filas crudas (con su nº de fila de Excel)
# ---------------------------------------------------------------------------
def _ref(hoja: str, fila: int | None = None, columna: str | None = None) -> str:
    """Ruta de error legible que cita hoja/fila/columna (p. ej. ``history.row5.units_sold``)."""
    partes = [hoja]
    if fila is not None:
        partes.append(f"row{fila}")
    if columna is not None:
        partes.append(columna)
    return ".".join(partes)


def _leer_hoja(
    wb: Any, hoja: HojaExcel, detalles: list[DetalleError]
) -> tuple[list[dict[str, Any]], list[int]]:
    """Lee una hoja: valida cabeceras y convierte cada celda al tipo del contrato.

    Devuelve ``(filas, nros_de_fila_excel)``. Acumula los problemas en ``detalles``
    (sin abortar a la primera) para reportar varios errores de una vez.
    """
    if hoja.nombre not in wb.sheetnames:
        detalles.append(DetalleError(field=_ref(hoja.nombre), problem="falta la hoja requerida."))
        return [], []

    ws = wb[hoja.nombre]
    filas_raw = list(ws.iter_rows(values_only=True))
    if not filas_raw:
        detalles.append(DetalleError(field=_ref(hoja.nombre), problem="la hoja está vacía."))
        return [], []

    # --- Cabecera (fila 1): mapea nombre de columna -> índice ---
    cabecera = filas_raw[0]
    columnas_def = {c.nombre: c for c in hoja.columnas}
    indice: dict[str, int] = {}
    for i, encabezado in enumerate(cabecera):
        if encabezado is None or str(encabezado).strip() == "":
            continue  # columnas en blanco al final: se ignoran
        nombre = str(encabezado).strip()
        if nombre not in columnas_def:
            detalles.append(
                DetalleError(
                    field=_ref(hoja.nombre, 1, nombre),
                    problem="columna desconocida (no está en el contrato).",
                )
            )
            continue
        indice[nombre] = i

    # Solo las columnas OBLIGATORIAS deben estar presentes; una opcional ausente se
    # trata como toda vacía (aplican los defaults del modelo).
    faltantes = [c.nombre for c in hoja.columnas if c.requerido and c.nombre not in indice]
    for nombre in faltantes:
        detalles.append(
            DetalleError(field=_ref(hoja.nombre, 1, nombre), problem="falta la columna requerida.")
        )
    if faltantes:
        # Sin las columnas obligatorias no se pueden leer las filas con fiabilidad.
        return [], []

    presentes = [c for c in hoja.columnas if c.nombre in indice]

    # --- Filas de datos (desde la fila 2) ---
    filas: list[dict[str, Any]] = []
    nros: list[int] = []
    for desplazamiento, fila_raw in enumerate(filas_raw[1:]):
        nro_excel = desplazamiento + 2  # fila 1 = cabecera
        # Salta filas completamente vacías (blancos al final, muy comunes).
        if all(_es_vacia(fila_raw[indice[c.nombre]]) for c in presentes):
            continue
        item: dict[str, Any] = {}
        for col in presentes:
            celda = fila_raw[indice[col.nombre]]
            if _es_vacia(celda):
                if col.requerido:
                    detalles.append(
                        DetalleError(
                            field=_ref(hoja.nombre, nro_excel, col.nombre),
                            problem="celda vacía en un campo obligatorio.",
                        )
                    )
                # Opcional vacío: se omite la clave → aplica el default del modelo.
                continue
            try:
                item[col.nombre] = convertir_celda(celda, col.conversor)
            except _CeldaInvalida as exc:
                letra = get_column_letter(indice[col.nombre] + 1)
                detalles.append(
                    DetalleError(
                        field=_ref(hoja.nombre, nro_excel, col.nombre),
                        problem=f"{exc} (celda {letra}{nro_excel}).",
                    )
                )
        filas.append(item)
        nros.append(nro_excel)

    if not filas:
        detalles.append(
            DetalleError(field=_ref(hoja.nombre), problem="no hay filas de datos.")
        )
    return filas, nros


# ---------------------------------------------------------------------------
# Mapeo de errores de Pydantic (sobre la petición) de vuelta a hoja/fila/columna
# ---------------------------------------------------------------------------
def _mapa_clave_a_hoja(plantilla: PlantillaDominio) -> tuple[dict[str, HojaExcel], HojaExcel | None]:
    """Indexa: clave de lista -> hoja, y devuelve la hoja escalar (parameters) si existe."""
    por_clave: dict[str, HojaExcel] = {}
    escalar: HojaExcel | None = None
    for hoja in plantilla.hojas:
        if hoja.es_lista and hoja.clave_peticion is not None:
            por_clave[hoja.clave_peticion] = hoja
        elif not hoja.es_lista:
            escalar = hoja
    return por_clave, escalar


def _detalle_desde_loc(
    loc: tuple[Any, ...],
    msg: str,
    por_clave: dict[str, HojaExcel],
    escalar: HojaExcel | None,
    nros: dict[str, list[int]],
    fila_escalar: int | None,
) -> DetalleError:
    """Traduce una ubicación de error Pydantic (``('history',3,'units_sold')``) a hoja/fila."""
    if loc and loc[0] in por_clave:
        hoja = por_clave[loc[0]]
        columna = str(loc[2]) if len(loc) >= 3 else None
        fila_excel = None
        if len(loc) >= 2 and isinstance(loc[1], int):
            lista = nros.get(hoja.nombre, [])
            if 0 <= loc[1] < len(lista):
                fila_excel = lista[loc[1]]
        return DetalleError(field=_ref(hoja.nombre, fila_excel, columna), problem=msg)
    if escalar is not None and loc:
        # Campo escalar (granularity/horizon): vive en la hoja de parámetros, fila de datos.
        return DetalleError(field=_ref(escalar.nombre, fila_escalar, str(loc[0])), problem=msg)
    return DetalleError(field=".".join(str(p) for p in loc) or "(excel)", problem=msg)


# ---------------------------------------------------------------------------
# API pública del lector
# ---------------------------------------------------------------------------
def leer_peticion(contenido: bytes, dominio: str) -> BaseModel:
    """Lee el ``.xlsx`` de un dominio y devuelve la **petición validada** del contrato.

    Convierte tipos explícitamente, arma la misma estructura que el JSON y la valida
    con el mismo modelo strict. Lanza ``ErrorExcel`` (→ HTTP 422) con hoja/fila/columna
    si el archivo es ilegible o no cumple el contrato.
    """
    plantilla = plantilla_de(dominio)
    try:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # archivo no es un .xlsx válido / corrupto
        raise ErrorExcel(
            [DetalleError(field="(archivo)", problem="no se pudo leer como Excel (.xlsx).")],
            "No se pudo leer el archivo Excel.",
        ) from exc

    try:
        detalles: list[DetalleError] = []
        peticion: dict[str, Any] = {}
        nros: dict[str, list[int]] = {}
        fila_escalar: int | None = None

        for hoja in plantilla.hojas:
            filas, fila_nros = _leer_hoja(wb, hoja, detalles)
            if hoja.es_lista and hoja.clave_peticion is not None:
                peticion[hoja.clave_peticion] = filas
                nros[hoja.nombre] = fila_nros
            else:
                # Parámetros escalares: una sola fila que se funde en la raíz.
                if filas:
                    peticion.update(filas[0])
                    fila_escalar = fila_nros[0]

        # Errores de estructura/conversión: cortar aquí con el reporte acumulado.
        if detalles:
            raise ErrorExcel(detalles)

        # Misma validación que el JSON, mapeando los fallos a hoja/fila/columna.
        modelo = MODELO_PETICION[dominio]
        try:
            return modelo.model_validate(peticion)
        except ValidationError as exc:
            por_clave, escalar = _mapa_clave_a_hoja(plantilla)
            mapeados = [
                _detalle_desde_loc(
                    tuple(e.get("loc", ())), str(e.get("msg", "")), por_clave, escalar, nros, fila_escalar
                )
                for e in exc.errors()
            ]
            raise ErrorExcel(mapeados) from exc
    finally:
        wb.close()
