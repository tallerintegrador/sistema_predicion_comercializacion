"""Generador de la plantilla ``.xlsx`` **a partir del contrato**.

La plantilla se construye desde la definición de `esquema_excel` (que a su vez deriva
de los modelos Pydantic), de modo que **plantilla y validación nunca se desincronizan**.
Cada archivo trae, por dominio:

- una hoja por bloque del contrato (`history`, `parameters`/`replenishment_params`/
  `inventory_status`) con **encabezados en inglés** y una **fila de ejemplo**;
- una hoja **`instructions`** en **español** que explica cada columna (tipo, si es
  obligatoria y los valores permitidos).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from spc.api.ingest.esquema_excel import (
    HOJA_INSTRUCCIONES,
    ColumnaExcel,
    HojaExcel,
    PlantillaDominio,
    plantilla_de,
)

# Estilos sobrios y consistentes para encabezados.
_AZUL = "2F6F8F"  # color_primary del proyecto
_GRIS = "EEF3F6"
_FONT_TITULO = Font(bold=True, size=13, color="1F4E5F")
_FONT_CABECERA = Font(bold=True, color="FFFFFF")
_FILL_CABECERA = PatternFill("solid", fgColor=_AZUL)
_FONT_SUBTITULO = Font(bold=True, color="1F4E5F")
_FILL_SUB = PatternFill("solid", fgColor=_GRIS)
_WRAP = Alignment(vertical="top", wrap_text=True)

_TIPO_HUMANO = {
    "date": "fecha (YYYY-MM-DD)",
    "id": "texto",
    "number": "número",
    "integer": "entero",
    "boolean": "booleano (true/false)",
    "enum": "texto",
}


def _tipo_humano(col: ColumnaExcel) -> str:
    """Nombre humano del tipo de una columna (para la hoja de instrucciones)."""
    base = _TIPO_HUMANO.get(col.conversor, "texto")
    if col.valores:
        return f"{base}: {', '.join(col.valores)}"
    return base


def _escribir_hoja_datos(ws: Worksheet, hoja: HojaExcel) -> None:
    """Escribe la cabecera (inglés) y las filas de ejemplo de una hoja de datos."""
    for j, col in enumerate(hoja.columnas, start=1):
        celda = ws.cell(row=1, column=j, value=col.nombre)
        celda.font = _FONT_CABECERA
        celda.fill = _FILL_CABECERA
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(col.nombre) + 4)
    for i, ejemplo in enumerate(hoja.ejemplos, start=2):
        for j, col in enumerate(hoja.columnas, start=1):
            ws.cell(row=i, column=j, value=ejemplo.get(col.nombre))
    ws.freeze_panes = "A2"


def _escribir_instrucciones(ws: Worksheet, plantilla: PlantillaDominio) -> None:
    """Escribe la hoja de ayuda en español: notas generales y tabla por columna."""
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 60

    fila = 1
    ws.cell(row=fila, column=1, value=f"Plantilla de {plantilla.titulo}").font = _FONT_TITULO
    fila += 2

    notas = [
        "Cómo usar esta plantilla:",
        "• Rellene una fila por observación/registro; no cambie los nombres de las columnas (en inglés).",
        "• Los encabezados están en inglés porque coinciden con el contrato de la API; estas notas, en español.",
        "• Fechas en formato ISO YYYY-MM-DD (p. ej. 2017-08-01).",
        "• Deje vacías las celdas de campos opcionales que no aplican; los obligatorios no pueden ir vacíos.",
        "• Los identificadores (store_id, product_id) se tratan como texto (el número 1 y el texto \"1\" son iguales).",
        "• Suba el archivo en el endpoint /{dominio}/excel; recibirá el mismo resultado que enviando JSON.",
    ]
    for n in notas:
        ws.cell(row=fila, column=1, value=n)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        fila += 1
    fila += 1

    for hoja in plantilla.hojas:
        ws.cell(row=fila, column=1, value=f"Hoja: {hoja.nombre}").font = _FONT_SUBTITULO
        fila += 1
        encabezados = ["Columna", "Tipo", "Obligatorio", "Valores permitidos", "Descripción"]
        for j, txt in enumerate(encabezados, start=1):
            c = ws.cell(row=fila, column=j, value=txt)
            c.font = _FONT_SUBTITULO
            c.fill = _FILL_SUB
        fila += 1
        for col in hoja.columnas:
            ws.cell(row=fila, column=1, value=col.nombre)
            ws.cell(row=fila, column=2, value=_TIPO_HUMANO.get(col.conversor, "texto"))
            ws.cell(row=fila, column=3, value="sí" if col.requerido else "no")
            ws.cell(row=fila, column=4, value=", ".join(col.valores) if col.valores else "—")
            d = ws.cell(row=fila, column=5, value=col.descripcion or "")
            d.alignment = _WRAP
            fila += 1
        fila += 1


def generar_workbook(dominio: str) -> Workbook:
    """Construye el ``Workbook`` de un dominio (instrucciones + hojas de datos)."""
    plantilla = plantilla_de(dominio)
    wb = Workbook()
    # La primera hoja por defecto se reutiliza como hoja de instrucciones.
    ws_instr = wb.active
    ws_instr.title = HOJA_INSTRUCCIONES
    _escribir_instrucciones(ws_instr, plantilla)
    for hoja in plantilla.hojas:
        ws = wb.create_sheet(title=hoja.nombre)
        _escribir_hoja_datos(ws, hoja)
    return wb


def generar_bytes(dominio: str) -> bytes:
    """Genera el ``.xlsx`` de un dominio en memoria y lo devuelve como ``bytes``."""
    wb = generar_workbook(dominio)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


__all__ = ["generar_workbook", "generar_bytes"]
