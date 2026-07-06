"""Generador de plantillas Excel ricas con instrucciones y datos de ejemplo.

Crea plantillas Excel con:
- Hoja 1: Instrucciones (columna, tipo, descripción, ejemplo)
- Hoja 2: Datos de ejemplo (5 sectores peruanos × módulo)
"""

import json
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from spc.catalogo.config import obtener_modulo

# ==============================================================================
# Datos de Ejemplo: 5 Sectores Peruanos × 3 Módulos
# ==============================================================================

EJEMPLOS_VENTAS = [
    # Bodega (mayorista)
    {"fecha": "2025-01-15", "id_tienda": "BODEGA-LI", "sku": "SKU-AZC-01", "categoria": "Abarrotes",
     "unidades_vendidas": 850, "precio_unitario": 2.50, "ingreso": 2125, "en_promocion": 0, "descuento_pct": 0,
     "metodo_pago": "Contado", "canal_venta": "Online", "es_fin_de_semana": 0, "dias_a_proximo_feriado": 10},
    # Botica
    {"fecha": "2025-01-15", "id_tienda": "BOTICA-JR", "sku": "SKU-VIT-05", "categoria": "Salud",
     "unidades_vendidas": 45, "precio_unitario": 125.00, "ingreso": 5625, "en_promocion": 1, "descuento_pct": 10,
     "metodo_pago": "Tarjeta", "canal_venta": "Tienda", "es_fin_de_semana": 1, "dias_a_proximo_feriado": 9},
    # Ferretería
    {"fecha": "2025-01-15", "id_tienda": "FERRE-NOR", "sku": "SKU-HER-12", "categoria": "Herramientas",
     "unidades_vendidas": 32, "precio_unitario": 450.00, "ingreso": 14400, "en_promocion": 0, "descuento_pct": 5,
     "metodo_pago": "Crédito", "canal_venta": "Tienda", "es_fin_de_semana": 0, "dias_a_proximo_feriado": 8},
    # Panadería
    {"fecha": "2025-01-15", "id_tienda": "PAN-AZNG", "sku": "SKU-PAN-03", "categoria": "Alimentos",
     "unidades_vendidas": 220, "precio_unitario": 8.50, "ingreso": 1870, "en_promocion": 1, "descuento_pct": 15,
     "metodo_pago": "Efectivo", "canal_venta": "Tienda", "es_fin_de_semana": 1, "dias_a_proximo_feriado": 7},
    # Restaurante
    {"fecha": "2025-01-15", "id_tienda": "REST-SJ", "sku": "SKU-PLA-08", "categoria": "Comidas",
     "unidades_vendidas": 156, "precio_unitario": 35.00, "ingreso": 5460, "en_promocion": 0, "descuento_pct": 0,
     "metodo_pago": "Tarjeta", "canal_venta": "Online", "es_fin_de_semana": 1, "dias_a_proximo_feriado": 6},
]

EJEMPLOS_COMPRAS = [
    # Bodega
    {"fecha_orden": "2025-01-10", "id_proveedor": "PROV-HUAY", "sku": "SKU-AZC-01", "categoria": "Abarrotes",
     "cantidad_pedida": 5000, "precio_unitario_compra": 1.80, "costo_total": 9000, "lead_time_dias": 3,
     "cantidad_recibida": 5000, "cumplimiento": 1.0, "metodo_pago": "Transferencia", "descuento_volumen": 0.08},
    # Botica
    {"fecha_orden": "2025-01-12", "id_proveedor": "PROV-FARM", "sku": "SKU-VIT-05", "categoria": "Salud",
     "cantidad_pedida": 200, "precio_unitario_compra": 95.00, "costo_total": 19000, "lead_time_dias": 5,
     "cantidad_recibida": 200, "cumplimiento": 1.0, "metodo_pago": "Crédito", "descuento_volumen": 0.05},
    # Ferretería
    {"fecha_orden": "2025-01-08", "id_proveedor": "PROV-IND", "sku": "SKU-HER-12", "categoria": "Herramientas",
     "cantidad_pedida": 150, "precio_unitario_compra": 350.00, "costo_total": 52500, "lead_time_dias": 7,
     "cantidad_recibida": 148, "cumplimiento": 0.987, "metodo_pago": "Transferencia", "descuento_volumen": 0.10},
    # Panadería
    {"fecha_orden": "2025-01-14", "id_proveedor": "PROV-HARINA", "sku": "SKU-PAN-03", "categoria": "Alimentos",
     "cantidad_pedida": 2000, "precio_unitario_compra": 5.50, "costo_total": 11000, "lead_time_dias": 1,
     "cantidad_recibida": 2000, "cumplimiento": 1.0, "metodo_pago": "Efectivo", "descuento_volumen": 0.03},
    # Restaurante
    {"fecha_orden": "2025-01-13", "id_proveedor": "PROV-REST", "sku": "SKU-PLA-08", "categoria": "Comidas",
     "cantidad_pedida": 800, "precio_unitario_compra": 15.00, "costo_total": 12000, "lead_time_dias": 2,
     "cantidad_recibida": 800, "cumplimiento": 1.0, "metodo_pago": "Transferencia", "descuento_volumen": 0.06},
]

EJEMPLOS_ALMACEN = [
    # Bodega
    {"fecha": "2025-01-15", "id_tienda": "BODEGA-LI", "sku": "SKU-AZC-01", "categoria": "Abarrotes",
     "stock_actual": 12000, "stock_minimo": 2000, "stock_maximo": 20000, "demanda_dia": 850,
     "demanda_diaria_promedio": 800, "dias_de_cobertura": 15.0, "rotacion": 6.5,
     "tiempo_reposicion_dias": 3, "zona_almacen": "Z-PRINCIPAL"},
    # Botica
    {"fecha": "2025-01-15", "id_tienda": "BOTICA-JR", "sku": "SKU-VIT-05", "categoria": "Salud",
     "stock_actual": 450, "stock_minimo": 100, "stock_maximo": 800, "demanda_dia": 45,
     "demanda_diaria_promedio": 42, "dias_de_cobertura": 10.0, "rotacion": 2.8,
     "tiempo_reposicion_dias": 5, "zona_almacen": "Z-REFRIGERACION"},
    # Ferretería
    {"fecha": "2025-01-15", "id_tienda": "FERRE-NOR", "sku": "SKU-HER-12", "categoria": "Herramientas",
     "stock_actual": 280, "stock_minimo": 50, "stock_maximo": 500, "demanda_dia": 32,
     "demanda_diaria_promedio": 35, "dias_de_cobertura": 8.0, "rotacion": 2.2,
     "tiempo_reposicion_dias": 7, "zona_almacen": "Z-HERRAMIENTAS"},
    # Panadería
    {"fecha": "2025-01-15", "id_tienda": "PAN-AZNG", "sku": "SKU-PAN-03", "categoria": "Alimentos",
     "stock_actual": 1500, "stock_minimo": 300, "stock_maximo": 2500, "demanda_dia": 220,
     "demanda_diaria_promedio": 210, "dias_de_cobertura": 7.0, "rotacion": 4.2,
     "tiempo_reposicion_dias": 1, "zona_almacen": "Z-FRESCO"},
    # Restaurante
    {"fecha": "2025-01-15", "id_tienda": "REST-SJ", "sku": "SKU-PLA-08", "categoria": "Comidas",
     "stock_actual": 250, "stock_minimo": 50, "stock_maximo": 400, "demanda_dia": 156,
     "demanda_diaria_promedio": 150, "dias_de_cobertura": 1.6, "rotacion": 8.5,
     "tiempo_reposicion_dias": 2, "zona_almacen": "Z-COCINA"},
]

# ==============================================================================
# Descripciones de Columnas
# ==============================================================================

DESCRIPCIONES_COLUMNAS = {
    "ventas": [
        ("fecha", "Fecha (YYYY-MM-DD)", "Día de la venta", "2025-01-15"),
        ("id_tienda", "ID Tienda", "Código único de la sucursal", "BODEGA-LI"),
        ("sku", "SKU", "Código del producto", "SKU-AZC-01"),
        ("categoria", "Categoría", "Tipo de producto", "Abarrotes"),
        ("unidades_vendidas", "Unidades Vendidas", "Cantidad vendida", "850"),
        ("precio_unitario", "Precio Unitario (S/)", "Precio por unidad", "2.50"),
        ("ingreso", "Ingreso (S/)", "Se calcula sola: unidades × precio", "2125"),
        ("en_promocion", "En Promoción (0/1)", "1 si está en promo, 0 si no", "0"),
        ("descuento_pct", "Descuento (%)", "% de descuento aplicado", "0"),
        ("metodo_pago", "Método de Pago", "Efectivo, Tarjeta, Crédito, etc.", "Contado"),
        ("canal_venta", "Canal de Venta", "Online o Tienda", "Online"),
        ("es_fin_de_semana", "¿Fin de Semana? (0/1)", "1 si es sábado/domingo", "0"),
        ("dias_a_proximo_feriado", "Días a Próximo Feriado", "Días hasta el feriado", "10"),
    ],
    "compras": [
        ("fecha_orden", "Fecha de Orden (YYYY-MM-DD)", "Día que se hizo el pedido", "2025-01-10"),
        ("id_proveedor", "ID Proveedor", "Código del proveedor", "PROV-HUAY"),
        ("sku", "SKU", "Código del producto", "SKU-AZC-01"),
        ("categoria", "Categoría", "Tipo de producto", "Abarrotes"),
        ("cantidad_pedida", "Cantidad Pedida", "Unidades solicitadas", "5000"),
        ("precio_unitario_compra", "Precio Compra (S/)", "Costo por unidad", "1.80"),
        ("costo_total", "Costo Total (S/)", "Se calcula sola: cantidad × precio", "9000"),
        ("lead_time_dias", "Lead Time (días)", "Días de entrega esperados", "3"),
        ("cantidad_recibida", "Cantidad Recibida", "Unidades que llegaron", "5000"),
        ("cumplimiento", "Cumplimiento (%)", "Ratio: cantidad recibida / pedida", "1.0"),
        ("metodo_pago", "Método de Pago", "Efectivo, Transferencia, Crédito", "Transferencia"),
        ("descuento_volumen", "Descuento por Volumen (%)", "% descuento por cantidad", "0.08"),
    ],
    "almacen": [
        ("fecha", "Fecha (YYYY-MM-DD)", "Día del inventario", "2025-01-15"),
        ("id_tienda", "ID Tienda", "Código único de la sucursal", "BODEGA-LI"),
        ("sku", "SKU", "Código del producto", "SKU-AZC-01"),
        ("categoria", "Categoría", "Tipo de producto", "Abarrotes"),
        ("stock_actual", "Stock Actual", "Unidades en almacén hoy", "12000"),
        ("stock_minimo", "Stock Mínimo", "Umbral mínimo de seguridad", "2000"),
        ("stock_maximo", "Stock Máximo", "Límite superior recomendado", "20000"),
        ("demanda_dia", "Demanda del Día", "Unidades vendidas/usadas hoy", "850"),
        ("demanda_diaria_promedio", "Demanda Diaria Promedio", "Promedio histórico", "800"),
        ("dias_de_cobertura", "Días de Cobertura", "Días que dura el stock", "15.0"),
        ("rotacion", "Rotación", "Velocidad de salida del stock", "6.5"),
        ("tiempo_reposicion_dias", "Tiempo Reposición (días)", "Días hasta nuevo pedido", "3"),
        ("zona_almacen", "Zona de Almacén", "Ubicación física", "Z-PRINCIPAL"),
    ],
}

# ==============================================================================
# Generador de Excel
# ==============================================================================

def generar_plantilla_excel(modulo: str) -> bytes:
    """Genera una plantilla Excel con instrucciones y datos de ejemplo.

    Hoja 1: Instrucciones (columna, tipo, descripción, ejemplo)
    Hoja 2: Datos de ejemplo (5 sectores peruanos)
    """
    # Validar módulo
    obtener_modulo(modulo)

    wb = Workbook()

    # Hoja 1: Instrucciones
    ws_instr = wb.active
    ws_instr.title = "Instrucciones"

    descripcion = DESCRIPCIONES_COLUMNAS.get(modulo, [])

    # Encabezados
    encabezados = ["Columna", "Tipo / Unidad", "Descripción", "Ejemplo"]
    for col, encabezado in enumerate(encabezados, 1):
        cell = ws_instr.cell(1, col)
        cell.value = encabezado
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filas de descripción
    for row, (col_name, tipo, desc, ejemplo) in enumerate(descripcion, 2):
        ws_instr.cell(row, 1).value = col_name
        ws_instr.cell(row, 2).value = tipo
        ws_instr.cell(row, 3).value = desc
        ws_instr.cell(row, 4).value = ejemplo

    # Ajustar ancho de columnas
    ws_instr.column_dimensions['A'].width = 25
    ws_instr.column_dimensions['B'].width = 25
    ws_instr.column_dimensions['C'].width = 40
    ws_instr.column_dimensions['D'].width = 20

    # Hoja 2: Datos de ejemplo
    ws_datos = wb.create_sheet("Datos de Ejemplo")

    # Seleccionar datos según módulo
    if modulo == "ventas":
        datos = EJEMPLOS_VENTAS
    elif modulo == "compras":
        datos = EJEMPLOS_COMPRAS
    else:  # almacen
        datos = EJEMPLOS_ALMACEN

    # Crear DataFrame
    df = pd.DataFrame(datos)

    # Escribir encabezados
    for col, column_title in enumerate(df.columns, 1):
        cell = ws_datos.cell(1, col)
        cell.value = column_title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="339933", end_color="339933", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Escribir datos
    for row, row_data in enumerate(df.values, 2):
        for col, value in enumerate(row_data, 1):
            cell = ws_datos.cell(row, col)
            cell.value = value
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")

    # Ajustar anchos
    for col in ws_datos.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_datos.column_dimensions[column].width = adjusted_width

    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==============================================================================
# Generador de plantilla JSON
# ==============================================================================

def generar_plantilla_json(modulo: str) -> bytes:
    """Genera una plantilla JSON con filas de ejemplo, lista para llenar y volver a subir.

    Estructura: {"_instrucciones": "...", "rows": [ ...filas de ejemplo... ]}.
    El endpoint de análisis acepta tanto este formato como un arreglo simple de filas.
    """
    obtener_modulo(modulo)  # valida el módulo

    datos = {
        "ventas": EJEMPLOS_VENTAS,
        "compras": EJEMPLOS_COMPRAS,
        "almacen": EJEMPLOS_ALMACEN,
    }[modulo]

    payload = {
        "_instrucciones": (
            "Reemplaza estas filas de ejemplo con tus datos reales. Mantén EXACTAMENTE los "
            "mismos nombres de columna. Puedes agregar todas las filas que necesites."
        ),
        "rows": datos,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
