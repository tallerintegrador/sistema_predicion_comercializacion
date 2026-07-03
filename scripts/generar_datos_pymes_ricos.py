"""Genera datasets de ejemplo **con señal causal real** para 2 rubros de PYME peruana.

A diferencia del motor sintético del sistema (que da señal débil), aquí se inyectan a mano
las relaciones que la auditoría pidió, para que el ML luzca y las gráficas sean informativas:

- Elasticidad precio→demanda NEGATIVA (más precio → menos unidades).
- Efecto de promoción, descuento, fin de semana y feriados peruanos reales.
- Canal de venta balanceado y dependiente de factores (categoría/precio).
- Cumplimiento de proveedor con varianza realista (σ≥0.10) y peor en proveedores lentos.
- Sobrestock (~12%), riesgo de quiebre (~15%) y reposición urgente reales.
- 6 zonas de almacén con perfiles distintos; varias tiendas/SKU/proveedores para clustering.

Reglas de consistencia obligatorias (se verifican al final):
  ingreso = unidades_vendidas × precio_unitario
  costo_total = cantidad_pedida × precio_unitario_compra
  cumplimiento = cantidad_recibida / cantidad_pedida   (cantidad_recibida = round(pedida × cumpl))
  dias_de_cobertura = stock_actual / demanda_diaria_promedio

Misma ventana temporal y granularidad para los tres módulos de cada sector (corrige A11).
Reproducible con ``--seed`` (default 2025). Escribe Excel (hoja Instrucciones + hoja Datos)
y JSON (lista plana) en ``ejemplos/datos_para_subir/<sector>/``.

Uso::

    ./venv/Scripts/python.exe scripts/generar_datos_pymes_ricos.py
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_RAIZ = Path(__file__).resolve().parent.parent
if str(_RAIZ / "src") not in sys.path:
    sys.path.insert(0, str(_RAIZ / "src"))

# ==============================================================================
# Calendario peruano
# ==============================================================================
VENTANA_INICIO = date(2023, 1, 1)
VENTANA_DIAS = 455  # ~15 meses → 2023-01-01 .. 2024-03-31 (mismo para los 3 módulos)

_FERIADOS_MD = [
    (1, 1), (4, 6), (4, 7), (5, 1), (6, 29), (7, 28), (7, 29), (8, 30),
    (10, 8), (11, 1), (12, 8), (12, 24), (12, 25), (12, 31),
]


def _feriados(anios: range) -> list[date]:
    fs = [date(a, m, d) for a in anios for (m, d) in _FERIADOS_MD]
    return sorted(f for f in fs)


def _fechas() -> list[date]:
    return [VENTANA_INICIO + timedelta(days=i) for i in range(VENTANA_DIAS)]


_FERIADOS = _feriados(range(2022, 2026))


def _dias_a_feriado(f: date) -> int:
    fut = [x for x in _FERIADOS if x >= f]
    return (fut[0] - f).days if fut else 99


# ==============================================================================
# Definición de rubros (identidad + parámetros causales)
# Producto: (nombre, categoria, precio_venta, elasticidad, demanda_base, zona)
# ==============================================================================
ZONAS = ["Recepción", "Góndola A", "Góndola B", "Refrigerados", "Bodega alta", "Trastienda"]

SECTORES: dict[str, dict] = {
    "minimarket_mi_barrio": {
        "titulo": "Minimarket / Bodega «Mi Barrio»",
        "tiendas": ["Bodega San Martín (SJL)", "Minimarket La Esquina (Comas)",
                    "Bodega Doña Rosa (VMT)", "Market El Ahorro (Ate)",
                    "Minimarket Los Olivos", "Bodega La Molina"],
        "proveedores": ["Distribuidora Alicorp", "Gloria S.A.", "Backus Depósito", "Molitalia",
                        "Comercial Andina", "Distribuidora Norte", "Mayorista Central",
                        "Proveedor La Victoria", "Nestlé Perú", "P&G Distribuidora",
                        "Laive Distribución", "San Fernando Mayorista"],
        # (nombre, categoria, precio, elasticidad, demanda_base)
        "productos": [
            ("Arroz Costeño 5kg", "Abarrotes", 22.0, 1.4, 40), ("Aceite Primor 1L", "Abarrotes", 9.5, 1.6, 35),
            ("Azúcar Rubia 1kg", "Abarrotes", 4.5, 1.3, 60), ("Fideos Don Vittorio 500g", "Abarrotes", 3.8, 1.5, 55),
            ("Atún Florida lata", "Abarrotes", 5.5, 1.7, 30), ("Leche Gloria tarro", "Lácteos", 4.2, 1.2, 80),
            ("Yogurt Gloria 1L", "Lácteos", 7.5, 1.8, 45), ("Queso Bonlé 500g", "Lácteos", 16.0, 2.0, 20),
            ("Inca Kola 1.5L", "Bebidas", 7.0, 1.9, 70), ("Agua San Luis 2.5L", "Bebidas", 5.0, 1.5, 65),
            ("Cerveza Pilsen 650ml", "Bebidas", 6.5, 2.2, 50), ("Detergente Bolívar 1kg", "Limpieza", 12.0, 1.6, 25),
            ("Papel higiénico x4", "Limpieza", 6.5, 1.4, 40), ("Galletas Soda Field", "Snacks", 2.5, 1.5, 90),
            ("Chocolate Sublime", "Snacks", 1.5, 1.8, 100), ("Shampoo H&S", "Cuidado personal", 18.0, 2.1, 15),
        ],
        "prob_promo": 0.22,
    },
    "ferreteria_construperu": {
        "titulo": "Ferretería «ConstruPerú»",
        "tiendas": ["Ferretería El Constructor (Ate)", "Ferretería Los Andes (VMT)",
                    "Ferretería Maestro Pérez (SJL)", "Ferretería La Unión (Comas)",
                    "Ferretería El Martillo (Callao)", "Ferretería Central (Lima)"],
        "proveedores": ["Cementos Pacasmayo", "Aceros Arequipa", "CPP Pinturas", "Distribuidora Eléctrica",
                        "Pavco Perú", "Ferretería Mayorista", "Importadora Sur", "Comercial Ferretera",
                        "Sodimac Distribución", "Promart Mayorista", "Celima Distribución", "Tekno Perú"],
        "productos": [
            ("Cemento Sol 42.5kg", "Construcción", 28.0, 1.1, 25), ("Fierro 1/2\" varilla", "Construcción", 32.0, 1.2, 18),
            ("Ladrillo King Kong", "Construcción", 0.9, 1.0, 120), ("Pintura Látex 1gal", "Pinturas", 45.0, 1.7, 10),
            ("Esmalte 1/4gal", "Pinturas", 22.0, 1.6, 12), ("Foco LED 9W", "Eléctrico", 8.0, 1.9, 30),
            ("Cable THW 14 (m)", "Eléctrico", 2.5, 1.5, 60), ("Interruptor simple", "Eléctrico", 6.0, 1.6, 22),
            ("Tubo PVC 1/2\" (m)", "Gasfitería", 6.0, 1.4, 35), ("Llave de paso 1/2\"", "Gasfitería", 14.0, 1.8, 14),
            ("Clavos 2\" (kg)", "Ferretería", 7.0, 1.3, 40), ("Cinta aislante", "Ferretería", 3.0, 1.5, 55),
            ("Brocha 3\"", "Pinturas", 9.0, 1.7, 20), ("Silicona transparente", "Ferretería", 12.0, 1.8, 16),
            ("Candado 40mm", "Seguridad", 18.0, 2.0, 12), ("Guantes de trabajo", "Seguridad", 10.0, 1.6, 24),
        ],
        "prob_promo": 0.12,
    },
}


# ==============================================================================
# Utilidades
# ==============================================================================
def _rng(seed: int, *tag: int) -> np.random.Generator:
    return np.random.default_rng([seed, *tag])


def _sig(x: float) -> float:
    return 1.0 / (1.0 + np.exp(-x))


def _perfil_skus(prods: list, rng: np.random.Generator) -> list[dict]:
    """Asigna a cada SKU su zona, rotación base y cobertura objetivo (por categoría)."""
    skus = []
    for i, (nombre, cat, precio, elast, dem) in enumerate(prods):
        zona = ZONAS[i % len(ZONAS)]
        rot_base = float(np.clip(rng.normal(2.5 - 0.02 * precio, 0.4), 0.2, 5.0))
        cob_obj = int(rng.integers(12, 30))  # días de cobertura objetivo (→ stock_maximo)
        repo = int(np.clip(round(rng.normal(6, 2)), 2, 15))
        skus.append(dict(id=f"SKU-{i+1:03d}", nombre=nombre, categoria=cat, precio=precio,
                         elasticidad=elast, demanda_base=dem, zona=zona, rotacion=rot_base,
                         cobertura=cob_obj, reposicion=repo))
    return skus


def _perfil_proveedores(provs: list, rng: np.random.Generator) -> list[dict]:
    """Cada proveedor: lead time medio, dispersión, cumplimiento base (peor si es lento), costo."""
    out = []
    for i, nombre in enumerate(provs):
        lead_medio = float(rng.uniform(3, 18))
        cumpl_base = float(np.clip(0.99 - 0.02 * (lead_medio - 3) + rng.normal(0, 0.03), 0.6, 0.99))
        out.append(dict(id=f"PROV-{i+1:02d}", nombre=nombre, lead=lead_medio,
                        lead_sigma=float(rng.uniform(1.0, 3.0)), cumpl=cumpl_base,
                        factor_costo=float(rng.uniform(0.85, 1.15)),
                        escala=float(rng.uniform(0.6, 1.6))))  # tamaño típico de pedido (para COM-R3)
    return out


# ==============================================================================
# VENTAS (≈10k filas: muestreo diario tienda×sku)
# ==============================================================================
def generar_ventas(sector: dict, skus: list[dict], seed: int) -> pd.DataFrame:
    rng = _rng(seed, 1)
    fechas = _fechas()
    tiendas = [f"T{t:02d}" for t in range(1, len(sector["tiendas"]) + 1)]
    mapa_t = {f"T{t:02d}": n for t, n in enumerate(sector["tiendas"], 1)}
    size_t = {t: float(rng.uniform(0.7, 1.4)) for t in tiendas}
    cat_online = {c: float(rng.uniform(-0.5, 0.6)) for c in {s["categoria"] for s in skus}}
    objetivo = 10000
    p_incl = objetivo / (len(fechas) * len(tiendas) * len(skus))

    filas = []
    for t in tiendas:
        for s in skus:
            r = _rng(seed, 2, int(t[1:]), int(s["id"][-3:]))
            for f in fechas:
                if r.random() > p_incl:
                    continue
                finde = int(f.weekday() >= 5)
                d_fer = _dias_a_feriado(f)
                f_finde = 1.0 + 0.25 * finde
                f_fer = 1.0 + 0.5 * max(0.0, (4 - d_fer) / 4.0)
                promo = int(r.random() < sector["prob_promo"])
                desc = round(float(r.uniform(8, 25)), 1) if promo else 0.0
                precio = round(s["precio"] * ((1 - desc / 100) if promo else float(r.uniform(0.97, 1.03))), 2)
                f_promo = 1.0 + (0.7 if promo else 0.0)
                f_precio = (precio / s["precio"]) ** (-s["elasticidad"])
                ruido = float(np.exp(r.normal(0, 0.12)))
                unidades = s["demanda_base"] * size_t[t] * f_finde * f_fer * f_promo * f_precio * ruido
                unidades = max(0, int(round(unidades)))
                if unidades == 0:
                    continue
                # canal dependiente de precio/categoría/finde (balanceado, multiclase)
                score_on = -0.2 + 0.9 * (s["precio"] > 12) + cat_online[s["categoria"]] + 0.3 * finde + r.normal(0, 0.4)
                p_on = _sig(score_on)
                u = r.random()
                canal = "online" if u < p_on * 0.7 else ("delivery" if u < p_on else "tienda")
                # B1: recargo de ticket por canal (delivery/online cobran más) → señal para VEN-R2.
                # No afecta las unidades (ya calculadas con el precio base), solo el precio cobrado.
                precio = round(precio * {"delivery": 1.12, "online": 1.06, "tienda": 1.0}[canal], 2)
                ingreso = round(unidades * precio, 2)
                metodo = str(r.choice(["efectivo", "yape_plin", "tarjeta", "transferencia"], p=[0.4, 0.3, 0.2, 0.1]))
                filas.append(dict(
                    fecha=f.isoformat(), id_tienda=mapa_t[t], sku=s["nombre"], categoria=s["categoria"],
                    unidades_vendidas=unidades, precio_unitario=precio, ingreso=ingreso,
                    en_promocion=promo, descuento_pct=desc, metodo_pago=metodo, canal_venta=canal,
                    es_fin_de_semana=finde, dias_a_proximo_feriado=int(d_fer),
                ))
    return pd.DataFrame(filas)


# ==============================================================================
# COMPRAS (≈10k órdenes: proveedor×sku, cada ~9 días en la ventana)
# ==============================================================================
def generar_compras(sector: dict, skus: list[dict], provs: list[dict], seed: int) -> pd.DataFrame:
    rc = _rng(seed, 33)
    cats = {s["categoria"] for s in skus}
    cat_lead = {c: float(rc.uniform(-2, 4)) for c in cats}
    cat_qty = {c: float(rc.uniform(0.6, 1.7)) for c in cats}  # tamaño típico por categoría (COM-R3)
    n_ord = max(1, round(10000 / (len(provs) * len(skus))))
    paso = max(1, VENTANA_DIAS // (n_ord + 1))
    filas = []
    for pv in provs:
        for s in skus:
            r = _rng(seed, 4, int(pv["id"][-2:]), int(s["id"][-3:]))
            # tamaño de pedido explicable por proveedor (escala) y categoría → señal para COM-R3
            base_qty = s["demanda_base"] * pv["escala"] * cat_qty[s["categoria"]] * float(r.uniform(4, 9))
            fase = int(r.integers(0, paso))
            for o in range(n_ord):
                f = VENTANA_INICIO + timedelta(days=min(VENTANA_DIAS - 1, fase + o * paso))
                lead = int(np.clip(round(r.normal(pv["lead"] + cat_lead[s["categoria"]], pv["lead_sigma"])), 1, 40))
                desc_vol = round(float(np.clip(r.normal(0.08, 0.04), 0.0, 0.2)), 3)
                precio = round(s["precio"] * 0.6 * pv["factor_costo"] * float(r.uniform(0.95, 1.06)), 2)
                cantidad = base_qty * (1 + 0.015 * lead) * (1 + 0.8 * desc_vol) * float(r.uniform(0.85, 1.15))
                cantidad = float(max(1, round(cantidad)))
                cumpl = float(np.clip(r.normal(pv["cumpl"] - 0.012 * (lead - pv["lead"]), 0.12), 0.55, 1.0))
                recibida = float(round(cantidad * cumpl))
                filas.append(dict(
                    fecha_orden=f.isoformat(), id_proveedor=pv["nombre"], sku=s["nombre"], categoria=s["categoria"],
                    cantidad_pedida=cantidad, precio_unitario_compra=precio,
                    costo_total=round(cantidad * precio, 2), lead_time_dias=lead,
                    cantidad_recibida=recibida, cumplimiento=round(recibida / cantidad, 4),
                    metodo_pago=str(r.choice(["contado", "credito_30", "credito_60"], p=[0.5, 0.35, 0.15])),
                    descuento_volumen=desc_vol,
                ))
    return pd.DataFrame(filas)


# ==============================================================================
# ALMACÉN (≈10k snapshots tienda×sku en fechas regulares)
# ==============================================================================
def generar_almacen(sector: dict, skus: list[dict], seed: int) -> pd.DataFrame:
    rng = _rng(seed, 5)
    tiendas = [f"T{t:02d}" for t in range(1, len(sector["tiendas"]) + 1)]
    mapa_t = {f"T{t:02d}": n for t, n in enumerate(sector["tiendas"], 1)}
    zona_dem = {z: float(rng.uniform(0.7, 1.4)) for z in ZONAS}
    zona_rot = {z: float(rng.uniform(0.7, 1.4)) for z in ZONAS}
    cat_dem = {s["categoria"]: float(rng.uniform(0.8, 1.3)) for s in skus}
    # B6: snapshots muestreados en TODA la ventana (misma densidad de fechas que ventas).
    all_fechas = _fechas()
    p_incl = 10000 / (len(all_fechas) * len(tiendas) * len(skus))

    filas = []
    for t in tiendas:
        size = float(rng.uniform(0.7, 1.4))
        for s in skus:
            r = _rng(seed, 6, int(t[1:]), int(s["id"][-3:]))
            dem_prom = s["demanda_base"] * size * cat_dem[s["categoria"]] * zona_dem[s["zona"]] * 0.3
            dem_prom = round(max(1.0, dem_prom), 2)
            rot = round(max(0.1, s["rotacion"] * zona_rot[s["zona"]] * float(r.uniform(0.85, 1.15))), 2)
            repo = s["reposicion"]
            stock_max = int(round(dem_prom * s["cobertura"]))
            # stock_mínimo = STOCK DE SEGURIDAD (1.5× demanda del lead time), distinto del
            # umbral de quiebre (ALM-C1) para diferenciar ALM-C2 "reposición urgente".
            stock_min = int(round(dem_prom * repo * 1.5))
            for f in all_fechas:
                if r.random() > p_incl:
                    continue
                finde = int(f.weekday() >= 5)
                demanda_dia = max(0, int(round(dem_prom * (1.15 if finde else 1.0) * float(np.exp(r.normal(0, 0.18))))))
                # cobertura objetivo con casos de quiebre (~15%) y sobrestock (~12%)
                u = r.random()
                if u < 0.15:
                    cov = float(r.uniform(0.2, 0.95)) * repo          # quiebre (ALM-C1)
                elif u < 0.27:
                    cov = s["cobertura"] + float(r.uniform(1, 12))    # sobrestock (ALM-C3)
                else:
                    cov = float(r.uniform(repo, s["cobertura"]))
                stock_actual = int(round(dem_prom * cov))
                dias_cob = round(stock_actual / dem_prom, 2)
                filas.append(dict(
                    fecha=f.isoformat(), id_tienda=mapa_t[t], sku=s["nombre"], categoria=s["categoria"],
                    stock_actual=stock_actual, stock_minimo=stock_min, stock_maximo=stock_max,
                    demanda_dia=demanda_dia, demanda_diaria_promedio=dem_prom, dias_de_cobertura=dias_cob,
                    rotacion=rot, tiempo_reposicion_dias=repo, zona_almacen=s["zona"],
                ))
    return pd.DataFrame(filas)


# ==============================================================================
# Escritura (Excel: Instrucciones + Datos; JSON: lista plana)
# ==============================================================================
_DESCR = {
    "fecha": "Fecha de la venta (AAAA-MM-DD)", "fecha_orden": "Fecha de la orden de compra",
    "id_tienda": "Nombre/código del local", "id_proveedor": "Nombre/código del proveedor",
    "sku": "Producto", "categoria": "Categoría del producto",
    "unidades_vendidas": "Unidades vendidas (entero)", "precio_unitario": "Precio de venta en S/",
    "ingreso": "= unidades_vendidas × precio_unitario", "en_promocion": "1 si estaba en promoción, 0 si no",
    "descuento_pct": "Descuento aplicado (%)", "metodo_pago": "Forma de pago",
    "canal_venta": "Canal (tienda/online/delivery)", "es_fin_de_semana": "1 si es sábado/domingo",
    "dias_a_proximo_feriado": "Días hasta el próximo feriado",
    "cantidad_pedida": "Unidades pedidas", "precio_unitario_compra": "Costo de compra en S/",
    "costo_total": "= cantidad_pedida × precio_unitario_compra", "lead_time_dias": "Días de entrega del proveedor",
    "cantidad_recibida": "Unidades efectivamente recibidas", "cumplimiento": "= cantidad_recibida / cantidad_pedida",
    "descuento_volumen": "Descuento por volumen (0..1)",
    "stock_actual": "Stock disponible hoy", "stock_minimo": "Stock mínimo de seguridad",
    "stock_maximo": "Stock máximo recomendado", "demanda_dia": "Demanda del día",
    "demanda_diaria_promedio": "Demanda diaria promedio", "dias_de_cobertura": "= stock_actual / demanda_diaria_promedio",
    "rotacion": "Índice de rotación", "tiempo_reposicion_dias": "Días para reponer", "zona_almacen": "Zona del almacén",
}


def _escribir_excel(df: pd.DataFrame, titulo: str, destino: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ins = wb.create_sheet("Instrucciones")
    ins.append(["SPC — Plantilla de datos", titulo])
    ins["A1"].font = Font(bold=True, size=13)
    ins.append([])
    ins.append(["Columna", "Qué es"])
    for c in ("A3", "B3"):
        ins[c].font = Font(bold=True)
        ins[c].fill = PatternFill("solid", fgColor="E2E8F0")
    for col in df.columns:
        ins.append([col, _DESCR.get(col, "")])
    ins.append([])
    ins.append(["Reemplaza la hoja «Datos» con tus registros (mismas columnas) y súbela en el Paso 3."])
    ins.column_dimensions["A"].width = 26
    ins.column_dimensions["B"].width = 52

    ws = wb.create_sheet("Datos")
    ws.append(list(df.columns))
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEAFE")
    for _, fila in df.iterrows():
        ws.append(list(fila.values))
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)


def _verificar(sector_slug: str, ventas: pd.DataFrame, compras: pd.DataFrame, almacen: pd.DataFrame) -> None:
    print(f"\n=== VERIFICACIÓN · {sector_slug} ===")
    for nombre, df, claves in [
        ("ventas", ventas, ["fecha", "id_tienda", "sku"]),
        ("compras", compras, ["fecha_orden", "id_proveedor", "sku"]),
        ("almacen", almacen, ["fecha", "id_tienda", "sku"]),
    ]:
        nulos = int(df.isna().sum().sum())
        dups = int(df.duplicated(subset=claves).sum())
        print(f"  {nombre:8} filas={len(df):6}  nulos={nulos}  duplicados={dups}")

    # Reglas de consistencia
    ok_ing = np.allclose(ventas["ingreso"], (ventas["unidades_vendidas"] * ventas["precio_unitario"]).round(2), atol=0.05)
    ok_cost = np.allclose(compras["costo_total"], (compras["cantidad_pedida"] * compras["precio_unitario_compra"]).round(2), atol=0.05)
    ok_cumpl = np.allclose(compras["cumplimiento"], (compras["cantidad_recibida"] / compras["cantidad_pedida"]).round(4), atol=0.002)
    ok_cob = np.allclose(almacen["dias_de_cobertura"], (almacen["stock_actual"] / almacen["demanda_diaria_promedio"]).round(2), atol=0.05)
    print(f"  identidades: ingreso={ok_ing} costo_total={ok_cost} cumplimiento={ok_cumpl} cobertura={ok_cob}")

    # Metas de balance/señal
    corr = ventas["unidades_vendidas"].corr(ventas["precio_unitario"])
    canal = ventas["canal_venta"].value_counts(normalize=True).round(2).to_dict()
    p75 = ventas["unidades_vendidas"].quantile(0.75)
    dia_pico = float((ventas["unidades_vendidas"] > p75).mean())
    sigma_cumpl = float(compras["cumplimiento"].std())
    sobre = float((almacen["stock_actual"] > almacen["stock_maximo"]).mean())
    quiebre = float((almacen["stock_actual"] < almacen["demanda_diaria_promedio"] * almacen["tiempo_reposicion_dias"]).mean())
    urgente = float((almacen["stock_actual"] < almacen["stock_minimo"]).mean())  # ALM-C2 (nueva regla)
    zonas = almacen["zona_almacen"].nunique()
    fechas_v = ventas["fecha"].nunique()
    fechas_c = compras["fecha_orden"].nunique()
    fechas_a = almacen["fecha"].nunique()
    print(f"  META corr(unidades,precio)={corr:+.2f} (objetivo: negativa)")
    print(f"  META canal={canal}  dia_pico={dia_pico:.0%}  sigma_cumplimiento={sigma_cumpl:.2f} (≥0.10)")
    print(f"  META sobrestock={sobre:.0%} (10-15%)  quiebre={quiebre:.0%}  reposicion={urgente:.0%}  zonas={zonas} (≥5)")
    print(f"  META fechas únicas (B6): ventas={fechas_v} compras={fechas_c} almacen={fechas_a} (deben ser similares)")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Genera datasets ricos con señal causal (2 rubros PYME peruana).")
    parser.add_argument("-o", "--out", default="ejemplos/datos_para_subir", help="Carpeta de salida.")
    parser.add_argument("--seed", type=int, default=2025)
    args = parser.parse_args(argv)
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except (AttributeError, ValueError):
        pass

    out = Path(args.out)
    for si, (slug, sector) in enumerate(SECTORES.items()):
        skus = _perfil_skus(sector["productos"], _rng(args.seed, 10, si))
        provs = _perfil_proveedores(sector["proveedores"], _rng(args.seed, 11, si))
        ventas = generar_ventas(sector, skus, args.seed)
        compras = generar_compras(sector, skus, provs, args.seed)
        almacen = generar_almacen(sector, skus, args.seed)

        carpeta = out / slug
        for nombre, df in [("ventas", ventas), ("compras", compras), ("almacen", almacen)]:
            (carpeta).mkdir(parents=True, exist_ok=True)
            (carpeta / f"{nombre}.json").write_text(
                json.dumps(df.to_dict("records"), ensure_ascii=False, indent=2), encoding="utf-8")
            _escribir_excel(df, sector["titulo"], carpeta / f"{nombre}.xlsx")
        print(f"[ok] {sector['titulo']} -> {carpeta}")
        _verificar(slug, ventas, compras, almacen)

    print("\nListo. Sube cualquiera en la app (Paso 3) o con POST /v3/{modulo}/archivo.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
