"""Generador de Excels sintéticos **conformes al contrato** con productos de PYME peruana.

Igual de conforme que ``generar_sinteticos.py`` (hojas y columnas se derivan de
``spc.api.ingest.esquema_excel``, la **misma** fuente que valida las subidas), pero el
vocabulario es de **pequeñas y medianas empresas peruanas**, escrito **en español**:
una bodega de abarrotes, una ferretería, una librería escolar, una distribuidora de
golosinas y un mayorista textil. Nada de bebidas, limpieza ni panadería.

Produce, en una carpeta nueva y **ordenada por dominio**:

- ``ventas/``   — pronóstico de demanda (hoja ``history``).
- ``compras/``  — reposición (``history`` + ``replenishment_params``).
- ``almacen/``  — riesgo de quiebre (``history`` + ``inventory_status``).
- ``invalidos/``— archivos rotos a propósito (deben dar **422**), variados.

Cada dominio trae tamaños **pequeño / mediano / grande**. Todo deriva de ``--seed``
(reproducible: misma semilla → mismos bytes). Solo depende de ``numpy``/``openpyxl``.

Uso::

    ./venv/Scripts/python.exe scripts/generar_sinteticos_peru.py --out data/synthetic_peru --seed 2025
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook

# Permite ejecutar el script directamente con el layout ``src/``.
_RAIZ = Path(__file__).resolve().parent.parent
if str(_RAIZ / "src") not in sys.path:
    sys.path.insert(0, str(_RAIZ / "src"))

from spc.api.ingest.esquema_excel import (  # noqa: E402
    HojaExcel,
    PlantillaDominio,
    plantilla_de,
)

# ---------------------------------------------------------------------------
# Catálogos reales de PYME peruana (en español). NADA de bebidas/limpieza/pan.
# Cada "giro" trae sus productos y los nombres de sus locales/sucursales.
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Giro:
    """Un rubro de PYME: sus productos, sus locales y la escala de su demanda."""

    nombre: str
    tiendas: tuple[str, ...]
    productos: tuple[str, ...]
    nivel: tuple[float, float]            # unidades/día por serie (rango base)
    amp_anual: tuple[float, float] = (0.10, 0.25)
    meses_pico: tuple[int, ...] = ()      # meses con repunte (p. ej. campaña escolar)
    pico_escolar: float = 0.0             # +% de demanda en meses_pico
    intermitencia: float = 0.0            # prob. de día sin venta


ABARROTES = Giro(
    nombre="Bodega de abarrotes",
    tiendas=("Bodega Doña Rosa", "Minimarket San Martín", "Bodega El Ahorro", "Market La Esquina"),
    productos=(
        "Arroz Costeño Superior 5kg",
        "Azúcar Rubia Cartavio 1kg",
        "Aceite Vegetal Primor 1L",
        "Fideos Don Vittorio Spaghetti 500g",
        "Atún Florida Filete en Aceite 170g",
        "Leche Gloria Evaporada 400g",
        "Lenteja Costeño 500g",
        "Quinua Perlada Wong 500g",
        "Avena Tres Ositos 170g",
        "Sal de Mesa Emsal 1kg",
        "Frijol Canario Costeño 500g",
        "Huevo de Gallina (kg)",
    ),
    nivel=(8.0, 120.0),
    intermitencia=0.02,
)

FERRETERIA = Giro(
    nombre="Ferretería",
    tiendas=("Ferretería El Constructor", "Ferretería Los Andes", "Ferretería Maestro Pérez"),
    productos=(
        "Foco LED Philips 9W",
        "Clavos de 2 pulgadas (kg)",
        "Cinta Aislante 3M Negra",
        "Pintura Látex Blanco Galón",
        "Candado Forte 40mm",
        "Brocha Tumi 3 pulgadas",
        "Silicona Transparente 280ml",
        "Cemento Sol Tipo I 42.5kg",
        "Tubo PVC 1/2 pulgada x3m",
        "Cerradura Travex Pomo",
    ),
    nivel=(1.0, 30.0),
    intermitencia=0.12,
)

LIBRERIA = Giro(
    nombre="Librería escolar",
    tiendas=("Librería Mundo Escolar", "Bazar y Librería Star", "Librería El Lápiz Feliz"),
    productos=(
        "Cuaderno Loro Cuadriculado 100h",
        "Lapicero Faber-Castell Trilux Azul",
        "Colores Artesco x12",
        "Cartulina Escolar A4 Blanca",
        "Plumón Artesco N°47 Negro",
        "Goma en Barra Stick 21g",
        "Regla Plástica Artesco 30cm",
        "Mochila Escolar Clásica",
        "Témpera Artesco x7 colores",
        "Corrector Líquido Vinifan",
    ),
    nivel=(2.0, 60.0),
    amp_anual=(0.20, 0.35),
    meses_pico=(2, 3),              # campaña escolar (febrero–marzo)
    pico_escolar=1.4,
    intermitencia=0.06,
)

GOLOSINAS = Giro(
    nombre="Distribuidora de golosinas",
    tiendas=(
        "Distribuidora La Económica",
        "Distribuidora Dulce Perú",
        "Almacén Gamarra Centro",
        "Distribuidora Golosinas del Sur",
    ),
    productos=(
        "Chocolate Sublime 30g",
        "Caramelo Cua Cua 12g",
        "Chizito Karinto 35g",
        "Papas Lays Clásica 40g",
        "Chocolate Triángulo Donofrio 30g",
        "Gomitas Mogul Frutas 40g",
        "Chupetín Globo Pop",
        "Wafer Chocman x6",
        "Caramelo Halls Mentol",
        "Chocolate Princesa Donofrio 32g",
    ),
    nivel=(30.0, 260.0),
    meses_pico=(7, 12),            # Fiestas Patrias y Navidad
    pico_escolar=0.6,
    intermitencia=0.0,
)

TEXTIL = Giro(
    nombre="Mayorista textil (Gamarra)",
    tiendas=("Mayorista Mesa Redonda", "Depósito Textil San Juan", "Galería Gamarra Stand 214"),
    productos=(
        "Polo Algodón Pima Talla M",
        "Medias Algodón Pack x3",
        "Chompa Lana Alpaca Talla L",
        "Buzo Deportivo Talla L",
        "Jean Clásico Talla 32",
        "Camisa Manga Larga Talla M",
        "Pijama Algodón Talla S",
        "Casaca Cortaviento Talla XL",
    ),
    nivel=(5.0, 90.0),
    amp_anual=(0.25, 0.40),
    meses_pico=(5, 7, 12),         # Día de la Madre, Fiestas Patrias, Navidad
    pico_escolar=0.8,
    intermitencia=0.08,
)


# ---------------------------------------------------------------------------
# Perfil de un archivo a generar
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Perfil:
    """Define un archivo: dominio, giro, forma (tiendas×productos×días) y tamaño."""

    nombre: str
    dominio: str          # "sales" | "purchases" | "inventory"
    giro: Giro
    n_tiendas: int
    n_productos: int
    n_dias: int
    fecha_inicio: str     # ISO YYYY-MM-DD
    tamano: str           # "pequeño" | "mediano" | "grande"
    descripcion: str
    promo_lambda: float = 0.4
    tendencia: tuple[float, float] = (-0.10, 0.35)
    ruido: tuple[float, float] = (0.08, 0.18)        # sigma relativo del ruido multiplicativo
    amp_semanal: tuple[float, float] = (0.30, 0.70)  # amplitud del patrón día-de-la-semana

    @property
    def filas_history(self) -> int:
        return self.n_tiendas * self.n_productos * self.n_dias

    @property
    def filas_extra(self) -> int:
        return self.n_tiendas * self.n_productos if self.dominio in ("purchases", "inventory") else 0


# Feriados peruanos relevantes (mes, día): activan event_active + un pico de demanda.
_FERIADOS_PE: frozenset[tuple[int, int]] = frozenset({
    (1, 1),    # Año Nuevo
    (5, 1),    # Día del Trabajo
    (6, 29),   # San Pedro y San Pablo
    (7, 28), (7, 29),  # Fiestas Patrias
    (8, 30),   # Santa Rosa de Lima
    (10, 8),   # Combate de Angamos
    (11, 1),   # Todos los Santos
    (12, 8),   # Inmaculada Concepción
    (12, 24), (12, 25),  # Navidad
    (12, 31),  # Fin de Año
})

# Repunte de fin de semana (Lun..Dom): típico del retail peruano.
_DOW_FACTOR = np.array([0.00, -0.04, -0.02, 0.03, 0.14, 0.30, 0.16])


# ---------------------------------------------------------------------------
# El MANIFIESTO: pequeño / mediano / grande por dominio + giros variados
# ---------------------------------------------------------------------------
PERFILES: tuple[Perfil, ...] = (
    # ---- VENTAS -----------------------------------------------------------
    Perfil(
        "ventas_bodega_pequeno", "sales", ABARROTES, 1, 3, 120, "2024-01-01", "pequeño",
        "Bodega chica: 1 local, 3 abarrotes de alta rotación, 4 meses de historia.",
        promo_lambda=0.5,
    ),
    Perfil(
        "ventas_libreria_mediano", "sales", LIBRERIA, 2, 6, 180, "2023-09-01", "mediano",
        "Librería: 2 locales, 6 útiles, con campaña escolar (febrero–marzo) marcada.",
        promo_lambda=0.7,
    ),
    Perfil(
        "ventas_golosinas_grande", "sales", GOLOSINAS, 4, 10, 240, "2023-06-01", "grande",
        "Distribuidora de golosinas: 4 almacenes × 10 productos × 240 días (9.600 filas).",
        promo_lambda=1.2,
    ),
    # ---- COMPRAS ----------------------------------------------------------
    Perfil(
        "compras_ferreteria_pequeno", "purchases", FERRETERIA, 1, 3, 100, "2024-02-01", "pequeño",
        "Ferretería chica: 3 productos de baja rotación + parámetros de reposición.",
        promo_lambda=0.2,
    ),
    Perfil(
        "compras_abarrotes_mediano", "purchases", ABARROTES, 2, 5, 160, "2023-10-01", "mediano",
        "Bodega mediana: 2 locales × 5 abarrotes + reposición (10 series).",
        promo_lambda=0.6,
    ),
    Perfil(
        "compras_distribuidora_grande", "purchases", GOLOSINAS, 4, 8, 220, "2023-06-01", "grande",
        "Distribuidora grande: 4 almacenes × 8 productos × 220 días + reposición (32 series).",
        promo_lambda=1.1,
    ),
    # ---- ALMACEN ----------------------------------------------------------
    Perfil(
        "almacen_libreria_pequeno", "inventory", LIBRERIA, 1, 3, 110, "2024-01-15", "pequeño",
        "Librería chica: 3 útiles + estado de inventario (lead time opcional).",
        promo_lambda=0.3,
    ),
    Perfil(
        "almacen_textil_mediano", "inventory", TEXTIL, 2, 6, 170, "2023-09-15", "mediano",
        "Mayorista textil: 2 stands × 6 prendas con estacionalidad fuerte (12 series).",
        promo_lambda=0.5,
    ),
    Perfil(
        "almacen_abarrotes_grande", "inventory", ABARROTES, 4, 10, 230, "2023-06-15", "grande",
        "Bodega/mayorista: 4 locales × 10 abarrotes × 230 días + inventario (40 series).",
        promo_lambda=1.0,
    ),
)


# ---------------------------------------------------------------------------
# Perfil especial para REENTRENAMIENTO por cliente (debe superar al congelado)
# ---------------------------------------------------------------------------
# Histórico rico de 2 años con estacionalidad semanal fuerte, tendencia clara,
# promociones/feriados marcados y ruido bajo: muy aprendible para el candidato por
# cliente. El congelado (entrenado con otra escala/productos) predice mal estos
# productos peruanos → el candidato lo supera en validación honesta (beats_frozen).
PERFIL_REENTRENAMIENTO = Perfil(
    "ventas_reentrenamiento_rico", "sales", ABARROTES, 4, 12, 730, "2023-01-01",
    "grande (reentrenamiento)",
    "Histórico rico de 2 años (48 series, ~35k filas) pensado para superar al modelo "
    "congelado al reentrenar por cliente.",
    promo_lambda=0.8,
    tendencia=(0.10, 0.30),
    ruido=(0.04, 0.07),          # poco ruido → señal aprendible
    amp_semanal=(0.45, 0.55),    # patrón semanal fuerte y estable
)


# ---------------------------------------------------------------------------
# Construcción de una serie de tiempo realista
# ---------------------------------------------------------------------------
def _rng(seed: int, *etiquetas: int) -> np.random.Generator:
    return np.random.default_rng([seed, *etiquetas])


def _entre(rng: np.random.Generator, rango: tuple[float, float]) -> float:
    return float(rng.uniform(rango[0], rango[1]))


def _fechas(inicio: str, n_dias: int) -> list[date]:
    d0 = date.fromisoformat(inicio)
    return [d0 + timedelta(days=i) for i in range(n_dias)]


def _serie(perfil: Perfil, fechas: list[date], rng: np.random.Generator) -> dict[str, np.ndarray]:
    """Demanda diaria de UNA serie: nivel × tendencia × semanal × anual × ruido + picos."""
    giro = perfil.giro
    n = len(fechas)
    t = np.arange(n, dtype=float)
    dow = np.array([f.weekday() for f in fechas])
    doy = np.array([f.timetuple().tm_yday for f in fechas], dtype=float)
    meses = np.array([f.month for f in fechas])

    nivel = _entre(rng, giro.nivel)
    tendencia = _entre(rng, perfil.tendencia)
    amp_sem = _entre(rng, perfil.amp_semanal)
    amp_anu = _entre(rng, giro.amp_anual)
    ruido = _entre(rng, perfil.ruido)
    fase = float(rng.uniform(0, 2 * np.pi))

    f_tendencia = 1.0 + tendencia * (t / max(1, n - 1))
    f_semanal = 1.0 + amp_sem * _DOW_FACTOR[dow]
    f_anual = 1.0 + amp_anu * np.sin(2 * np.pi * doy / 365.25 + fase)

    # Repunte de campaña (escolar / Fiestas Patrias / Navidad, según el giro).
    f_campana = np.ones(n)
    if giro.meses_pico and giro.pico_escolar:
        en_campana = np.isin(meses, np.array(giro.meses_pico))
        f_campana = np.where(en_campana, 1.0 + giro.pico_escolar, 1.0)

    promo = rng.poisson(perfil.promo_lambda, n).astype(int)
    promo_boost = _entre(rng, (0.10, 0.30))  # % extra por ítem en promo
    feriado = np.array([(f.month, f.day) in _FERIADOS_PE for f in fechas])
    factor_ruido = 1.0 + rng.normal(0.0, ruido, n)

    demanda = nivel * f_tendencia * f_semanal * f_anual * f_campana * factor_ruido
    demanda = demanda * (1.0 + promo_boost * np.minimum(promo, 3))
    demanda = demanda * np.where(feriado, _entre(rng, (1.3, 1.9)), 1.0)
    demanda = np.maximum(0.0, demanda)

    if giro.intermitencia > 0:
        cero = rng.random(n) < giro.intermitencia
        demanda = np.where(cero, 0.0, demanda)

    # transactions = flujo de clientes, correlacionado con la demanda; null si no hubo venta.
    ratio = float(rng.uniform(0.7, 1.4))
    transacciones = np.maximum(0.0, demanda * ratio + rng.normal(0, 2, n))

    return {"demanda": demanda, "promo": promo, "transacciones": transacciones, "feriado": feriado}


def _filas_history(perfil: Perfil, seed: int, idx: int) -> tuple[list[dict[str, Any]], dict[tuple[str, str], float]]:
    """Filas del bloque ``history`` y la demanda media reciente por serie."""
    fechas = _fechas(perfil.fecha_inicio, perfil.n_dias)
    tiendas = perfil.giro.tiendas[: perfil.n_tiendas]
    productos = perfil.giro.productos[: perfil.n_productos]
    filas: list[dict[str, Any]] = []
    media_reciente: dict[tuple[str, str], float] = {}

    for s, store_id in enumerate(tiendas):
        for p, product_id in enumerate(productos):
            rng = _rng(seed, idx, s, p)
            serie = _serie(perfil, fechas, rng)
            demanda, promo, trans, feriado = (
                serie["demanda"], serie["promo"], serie["transacciones"], serie["feriado"]
            )
            for i, f in enumerate(fechas):
                u = int(round(float(demanda[i])))  # unidades enteras (se vende por pieza)
                filas.append({
                    "date": f,
                    "store_id": store_id,
                    "product_id": product_id,
                    "units_sold": u,
                    "on_promotion": int(promo[i]),
                    "transactions": (None if u == 0 else int(round(float(trans[i])))),
                    "event_active": (True if bool(feriado[i]) else None),
                })
            ventana = demanda[-30:] if len(demanda) >= 30 else demanda
            media_reciente[(store_id, product_id)] = float(np.mean(ventana)) if len(ventana) else 0.0

    return filas, media_reciente


def _filas_dominio(perfil: Perfil, media_reciente: dict[tuple[str, str], float], seed: int, idx: int) -> list[dict[str, Any]]:
    """Filas de ``replenishment_params`` (compras) o ``inventory_status`` (almacén)."""
    if perfil.dominio not in ("purchases", "inventory"):
        return []
    rng = _rng(seed, idx, 9999)
    filas: list[dict[str, Any]] = []
    for (store_id, product_id), media in media_reciente.items():
        current_stock = round(max(0.0, media * float(rng.uniform(2.0, 7.0))), 0)
        lead = int(rng.integers(1, 12))  # 1..11 (> 0)
        if perfil.dominio == "purchases":
            filas.append({
                "store_id": store_id,
                "product_id": product_id,
                "current_stock": current_stock,
                "lead_time_days": lead,
                "target_coverage_days": int(rng.integers(3, 21)),  # 3..20 (> 0)
            })
        else:  # inventory: lead_time_days opcional → ~25% vacío.
            filas.append({
                "store_id": store_id,
                "product_id": product_id,
                "current_stock": current_stock,
                "lead_time_days": (None if rng.random() < 0.25 else lead),
            })
    return filas


# ---------------------------------------------------------------------------
# Escritura del .xlsx con las hojas/columnas canónicas del contrato
# ---------------------------------------------------------------------------
def _escribir(plantilla: PlantillaDominio, filas_por_hoja: dict[str, list[dict[str, Any]]], destino: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for hoja in plantilla.hojas:
        ws = wb.create_sheet(title=hoja.nombre)
        columnas = [c.nombre for c in hoja.columnas]
        ws.append(columnas)
        for fila in filas_por_hoja.get(hoja.nombre, []):
            ws.append([fila.get(c) for c in columnas])
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)


def _hoja_lista(plantilla: PlantillaDominio) -> HojaExcel:
    for hoja in plantilla.hojas:
        if hoja.es_lista and hoja.nombre != "history":
            return hoja
    raise KeyError("sin hoja-lista de dominio")


_SUBCARPETA = {"sales": "ventas", "purchases": "compras", "inventory": "almacen"}


def generar_archivo(perfil: Perfil, seed: int, idx: int, out_dir: Path) -> Path:
    plantilla = plantilla_de(perfil.dominio)
    filas_history, media_reciente = _filas_history(perfil, seed, idx)
    filas_por_hoja: dict[str, list[dict[str, Any]]] = {"history": filas_history}
    if perfil.dominio != "sales":
        hoja = _hoja_lista(plantilla)
        filas_por_hoja[hoja.nombre] = _filas_dominio(perfil, media_reciente, seed, idx)
    destino = out_dir / _SUBCARPETA[perfil.dominio] / f"{perfil.nombre}.xlsx"
    _escribir(plantilla, filas_por_hoja, destino)
    return destino


def generar_reentrenamiento(seed: int, out_dir: Path) -> Path:
    """Genera el Excel rico de SALES para reentrenamiento (+ su LEEME local).

    Va en su PROPIA subcarpeta (``reentrenamiento/``) porque su uso es distinto al del
    resto: alimenta ``POST /training/sales/excel``, no la predicción. Es solo hoja
    ``history`` (la plantilla de Ventas, ADR-0022).
    """
    p = PERFIL_REENTRENAMIENTO
    plantilla = plantilla_de(p.dominio)
    filas_history, _ = _filas_history(p, seed, 1000)  # idx alto: no choca con PERFILES
    destino = out_dir / "reentrenamiento" / f"{p.nombre}.xlsx"
    _escribir(plantilla, {"history": filas_history}, destino)

    leeme = out_dir / "reentrenamiento" / "LEEME.md"
    leeme.write_text(
        "\n".join([
            "# Datos para REENTRENAMIENTO por cliente (superan al modelo congelado)",
            "",
            f"`{p.nombre}.xlsx` — histórico de **SALES** (solo hoja `history`, la plantilla de "
            "Ventas) pensado para la función de **entrenamiento por cliente** "
            "(`POST /training/sales/excel`), no para la predicción.",
            "",
            f"- **{p.n_tiendas} locales × {p.n_productos} abarrotes = {p.n_tiendas * p.n_productos} "
            f"series**, **{p.n_dias} días** (2023-01-01 .. 2024-12-30) ≈ "
            f"**{p.filas_history:,} filas**.",
            "- Estacionalidad semanal fuerte, tendencia clara, promociones y feriados peruanos "
            "marcados, **ruido bajo** → un histórico **aprendible**.",
            "- Supera con holgura el gate de suficiencia (≥ 60 días / 120 filas / 1 serie) y lleva "
            "la ventana de validación temporal al **máximo (16 días, igual que el congelado)**.",
            "",
            "## Por qué supera al modelo congelado",
            "",
            "El congelado se entrenó con **otra escala y otros productos** (familias de "
            "Corporación Favorita, miles de unidades). Sobre estos productos peruanos (unidades "
            "en decenas/centenas) pronostica en una escala equivocada → **WAPE enorme**. El "
            "candidato, entrenado con los datos del propio cliente, aprende la escala real y le "
            "gana en la **misma ventana de validación honesta** → `beats_frozen = true`.",
            "",
            "## Uso",
            "",
            "```bash",
            "curl -F \"file=@reentrenamiento/" + p.nombre + ".xlsx\" \\",
            "     -H \"X-Client-Id: pyme-peru-demo\" \\",
            "     \"http://localhost:8000/training/sales/excel?source=excel\"",
            "# -> 202 {job_id};  luego  GET /training/jobs/{job_id}/result  -> comparación + veredicto",
            "```",
            "",
            "> Nota: «superar al congelado» (`beats_frozen`) es la condición que pide este archivo. "
            "La **adopción** además exige ganarle a un baseline ingenuo (media móvil de 7 días); "
            "el veredicto honesto (`adopted` / `not_adopted`) lo da el propio experimento.",
            "",
        ]),
        encoding="utf-8",
    )
    return destino


# ---------------------------------------------------------------------------
# Archivos rotos a propósito (deben dar 422), variados por dominio y por regla
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Invalido:
    nombre: str
    dominio: str
    motivo: str


INVALIDOS: tuple[Invalido, ...] = (
    Invalido("inv_ventas_texto_en_units_sold", "sales",
             "Texto ('agotado') en la columna numérica units_sold."),
    Invalido("inv_ventas_falta_product_id", "sales",
             "Falta la columna obligatoria product_id."),
    Invalido("inv_ventas_store_id_vacio", "sales",
             "Celda vacía en el campo obligatorio store_id."),
    Invalido("inv_ventas_fecha_invalida", "sales",
             "Fecha no válida ('30/02/2024') en la columna date."),
    Invalido("inv_ventas_units_negativo", "sales",
             "Valor negativo (-5) en units_sold (debe ser ≥ 0)."),
    Invalido("inv_compras_lead_time_cero", "purchases",
             "lead_time_days = 0 en replenishment_params (debe ser > 0)."),
)


def _base_valida(dominio: str, seed: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], HojaExcel | None]:
    """history pequeño y válido (+ hoja de dominio si aplica) para luego corromperlo."""
    plantilla = plantilla_de(dominio)
    perfil = Perfil("_base", dominio, ABARROTES, 1, 2, 25, "2024-01-01", "pequeño", "base")
    filas_history, media = _filas_history(perfil, seed, 555)
    hoja_dom = None
    filas_dom: list[dict[str, Any]] = []
    if dominio != "sales":
        hoja_dom = _hoja_lista(plantilla)
        filas_dom = _filas_dominio(perfil, media, seed, 555)
    return filas_history, filas_dom, hoja_dom


def generar_invalidos(seed: int, out_dir: Path) -> list[tuple[str, Path]]:
    resultados: list[tuple[str, Path]] = []
    for inv in INVALIDOS:
        plantilla = plantilla_de(inv.dominio)
        cols_history = [c.nombre for c in plantilla.hojas[0].columnas]
        filas_h, filas_dom, hoja_dom = _base_valida(inv.dominio, seed)
        filas_h = [dict(f) for f in filas_h]
        filas_dom = [dict(f) for f in filas_dom]
        cols_h = list(cols_history)

        if inv.nombre == "inv_ventas_texto_en_units_sold":
            filas_h[3]["units_sold"] = "agotado"
        elif inv.nombre == "inv_ventas_falta_product_id":
            cols_h = [c for c in cols_h if c != "product_id"]
        elif inv.nombre == "inv_ventas_store_id_vacio":
            filas_h[2]["store_id"] = None
        elif inv.nombre == "inv_ventas_fecha_invalida":
            filas_h[4]["date"] = "30/02/2024"
        elif inv.nombre == "inv_ventas_units_negativo":
            filas_h[5]["units_sold"] = -5
        elif inv.nombre == "inv_compras_lead_time_cero":
            filas_dom[0]["lead_time_days"] = 0

        wb = Workbook()
        wb.remove(wb.active)
        ws_h = wb.create_sheet(title="history")
        ws_h.append(cols_h)
        for f in filas_h:
            ws_h.append([f.get(c) for c in cols_h])
        if hoja_dom is not None:
            cols_d = [c.nombre for c in hoja_dom.columnas]
            ws_d = wb.create_sheet(title=hoja_dom.nombre)
            ws_d.append(cols_d)
            for f in filas_dom:
                ws_d.append([f.get(c) for c in cols_d])

        destino = out_dir / "invalidos" / f"{inv.nombre}.xlsx"
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)
        resultados.append((inv.motivo, destino))
    return resultados


# ---------------------------------------------------------------------------
# LEEME.md con el inventario de archivos (en español)
# ---------------------------------------------------------------------------
def escribir_leeme(out_dir: Path, seed: int) -> Path:
    ruta = out_dir / "LEEME.md"
    lineas = [
        "# Datos sintéticos — PYME peruana (productos reales, en español)",
        "",
        f"Generados con `scripts/generar_sinteticos_peru.py --seed {seed}` "
        "(reproducible: misma semilla → mismos archivos).",
        "",
        "Respetan EXACTAMENTE las plantillas del sistema (hojas y columnas en inglés, "
        "derivadas del contrato). Cada dominio trae tamaños **pequeño / mediano / grande**.",
        "",
        "## Ventas (`ventas/`) — hoja `history`",
        "",
        "| Archivo | Giro | Tiendas×Productos×Días | Filas | Tamaño |",
        "|---|---|---|---:|---|",
    ]
    for p in PERFILES:
        if p.dominio == "sales":
            lineas.append(
                f"| `{_SUBCARPETA[p.dominio]}/{p.nombre}.xlsx` | {p.giro.nombre} | "
                f"{p.n_tiendas}×{p.n_productos}×{p.n_dias} | {p.filas_history:,} | {p.tamano} |"
            )
    lineas += [
        "",
        "## Compras (`compras/`) — `history` + `replenishment_params`",
        "",
        "| Archivo | Giro | Tiendas×Productos×Días | Filas (+series) | Tamaño |",
        "|---|---|---|---:|---|",
    ]
    for p in PERFILES:
        if p.dominio == "purchases":
            lineas.append(
                f"| `{_SUBCARPETA[p.dominio]}/{p.nombre}.xlsx` | {p.giro.nombre} | "
                f"{p.n_tiendas}×{p.n_productos}×{p.n_dias} | {p.filas_history:,} (+{p.filas_extra}) | {p.tamano} |"
            )
    lineas += [
        "",
        "## Almacén (`almacen/`) — `history` + `inventory_status`",
        "",
        "| Archivo | Giro | Tiendas×Productos×Días | Filas (+series) | Tamaño |",
        "|---|---|---|---:|---|",
    ]
    for p in PERFILES:
        if p.dominio == "inventory":
            lineas.append(
                f"| `{_SUBCARPETA[p.dominio]}/{p.nombre}.xlsx` | {p.giro.nombre} | "
                f"{p.n_tiendas}×{p.n_productos}×{p.n_dias} | {p.filas_history:,} (+{p.filas_extra}) | {p.tamano} |"
            )
    lineas += [
        "",
        "## Inválidos (`invalidos/`) — deben dar **422** con un error legible",
        "",
        "| Archivo | Dominio | Qué tiene mal |",
        "|---|---|---|",
    ]
    for inv in INVALIDOS:
        lineas.append(f"| `invalidos/{inv.nombre}.xlsx` | {inv.dominio} | {inv.motivo} |")
    lineas.append("")
    ruta.write_text("\n".join(lineas), encoding="utf-8")
    return ruta


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Genera Excels sintéticos de PYME peruana, conformes al contrato SPC.")
    parser.add_argument("--out", default="data/synthetic_peru", help="Carpeta de salida (default: data/synthetic_peru).")
    parser.add_argument("--seed", type=int, default=2025, help="Semilla global (default: 2025).")
    parser.add_argument("--sin-invalidos", action="store_true", help="No generar los archivos rotos.")
    parser.add_argument("--solo-reentrenamiento", action="store_true",
                        help="Generar SOLO el Excel rico de reentrenamiento (no toca el resto).")
    args = parser.parse_args(argv)

    # La consola de Windows usa cp1252; forzamos UTF-8 para imprimir → ≥ × sin fallar.
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except (AttributeError, ValueError):
        pass

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Atajo: solo el archivo de reentrenamiento, sin tocar los demás.
    if args.solo_reentrenamiento:
        print(f"Generando SOLO reentrenamiento en {out_dir} con semilla {args.seed}\n")
        ruta = generar_reentrenamiento(args.seed, out_dir)
        mb = ruta.stat().st_size / (1024 * 1024)
        print(f"  [ok] reentrenamiento/{ruta.name:34s} "
              f"{PERFIL_REENTRENAMIENTO.filas_history:>6d} filas  {mb:5.2f} MB")
        print("\nListo.")
        return 0

    # Guarda: un perfil no puede pedir más tiendas/productos de los que tiene su giro
    # (si no, el slice se recortaría en silencio y los conteos quedarían deshonestos).
    for p in (*PERFILES, PERFIL_REENTRENAMIENTO):
        if p.n_tiendas > len(p.giro.tiendas) or p.n_productos > len(p.giro.productos):
            raise SystemExit(
                f"Perfil '{p.nombre}': pide {p.n_tiendas}×{p.n_productos} pero el giro "
                f"'{p.giro.nombre}' solo tiene {len(p.giro.tiendas)} tiendas × "
                f"{len(p.giro.productos)} productos."
            )

    print(f"Generando en {out_dir} con semilla {args.seed}\n")

    for idx, perfil in enumerate(PERFILES):
        ruta = generar_archivo(perfil, args.seed, idx, out_dir)
        mb = ruta.stat().st_size / (1024 * 1024)
        print(f"  [ok] {_SUBCARPETA[perfil.dominio]}/{ruta.name:36s} "
              f"{perfil.filas_history:>6d} filas  {perfil.tamano:8s}  {mb:5.2f} MB")

    if not args.sin_invalidos:
        print("\nArchivos rotos a propósito (para demostrar el manejo de errores → 422):")
        for motivo, ruta in generar_invalidos(args.seed, out_dir):
            print(f"  [x] invalidos/{ruta.name:34s} -> {motivo}")

    print("\nReentrenamiento por cliente (supera al modelo congelado):")
    ruta = generar_reentrenamiento(args.seed, out_dir)
    print(f"  [ok] reentrenamiento/{ruta.name:34s} {PERFIL_REENTRENAMIENTO.filas_history:>6d} filas")

    leeme = escribir_leeme(out_dir, args.seed)
    print(f"\n  [ok] {leeme.relative_to(out_dir)} (inventario de archivos)")
    print("\nListo.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
