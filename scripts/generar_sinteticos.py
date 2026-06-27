"""Generador de Excels sintéticos **conformes al contrato** del sistema SPC.

Produce archivos `.xlsx` ya **llenos**, realistas y reproducibles, que respetan
EXACTAMENTE las plantillas que el sistema provee por dominio
(`GET /{sales,purchases,inventory}/template`). Sirven para evaluar:

- el **modo en línea vs lote** (la frontera la decide ``len(history)`` contra
  ``SPC_ONLINE_MAX_ROWS``, default 2000), y
- el **reentrenamiento por cliente** bajo demanda.

Principios:

- **Conformidad.** Las hojas y los nombres de columna NO se escriben a mano: se
  derivan de ``spc.api.ingest.esquema_excel.PLANTILLAS`` (la **misma** fuente que el
  sistema usa para generar y validar las plantillas). Si el contrato cambia, los
  archivos lo siguen solo.
- **Realismo.** Cada serie ``(store_id, product_id)`` se construye con
  ``nivel_base × tendencia × estacionalidad_semanal × estacionalidad_anual +
  picos_por_promoción + eventos + ruido``, con intermitencia (ceros) opcional. No es
  ruido aleatorio plano.
- **Reproducibilidad.** Todo deriva de una única semilla (``--seed``): mismas
  entradas → mismos bytes. Solo depende de ``numpy``/``pandas``/``openpyxl``.
- **Control del modo.** El nº de filas (``tiendas × productos × días``) se elige para
  caer exacto en el modo previsto, **incluida la frontera 2000/2001**.

Uso::

    python scripts/generar_sinteticos.py --out data/sinteticos --seed 42

El validador (`scripts/validar_sinteticos.py`) comprueba después, archivo por archivo,
que (a) pasan la validación strict y (b) caen en el modo esperado (200 / 202).
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook

# Permite ejecutar el script directamente (`python scripts/...`) con layout `src/`.
_RAIZ = Path(__file__).resolve().parent.parent
if str(_RAIZ / "src") not in sys.path:
    sys.path.insert(0, str(_RAIZ / "src"))

from spc.api.ingest.esquema_excel import HojaExcel, PlantillaDominio, plantilla_de  # noqa: E402

# ---------------------------------------------------------------------------
# Vocabulario realista (familias reales del dataset Corporación Favorita).
# Las dos primeras (BEVERAGES, GROCERY) son las que conoce el artefacto diminuto
# de pruebas; van primero para que los archivos EN LÍNEA pronostiquen sin degradar.
# ---------------------------------------------------------------------------
FAMILIAS: tuple[str, ...] = (
    "BEVERAGES", "GROCERY", "PRODUCE", "CLEANING", "DAIRY",
    "MEATS", "BREAD/BAKERY", "POULTRY", "DELI", "EGGS",
    "FROZEN FOODS", "HOME CARE", "PERSONAL CARE", "SEAFOOD", "LIQUOR,WINE,BEER",
    "PREPARED FOODS", "BEAUTY", "HARDWARE", "AUTOMOTIVE", "LAWN AND GARDEN",
    "PET SUPPLIES", "SCHOOL AND OFFICE SUPPLIES", "MAGAZINES", "PLAYERS AND ELECTRONICS", "CELEBRATION",
)

# Efecto de día de la semana (Lun..Dom): retail con repunte de fin de semana.
_DOW_FACTOR = np.array([0.00, -0.05, -0.03, 0.02, 0.12, 0.28, 0.18])

# Días marcados como evento/feriado (mes, día): activan event_active=True + un pico.
_EVENTOS: frozenset[tuple[int, int]] = frozenset(
    {(1, 1), (5, 1), (8, 10), (11, 2), (12, 24), (12, 25), (12, 31)}
)


# ---------------------------------------------------------------------------
# Perfil de un archivo del manifiesto
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Perfil:
    """Define un archivo a generar: forma, fechas, patrón de demanda y modo esperado."""

    nombre: str
    dominio: str  # "sales" | "purchases" | "inventory"
    n_tiendas: int
    n_productos: int
    n_dias: int
    fecha_inicio: str  # ISO YYYY-MM-DD
    modo_esperado: str  # "online" | "batch"
    uso: str
    descripcion: str
    granularity: str | None = None  # solo sales
    horizon: int | None = None      # solo sales
    # Rangos del patrón de demanda (se sortea un valor por serie dentro del rango).
    nivel: tuple[float, float] = (200.0, 800.0)
    tendencia: tuple[float, float] = (-0.10, 0.40)      # cambio relativo a lo largo de la serie
    amp_semanal: tuple[float, float] = (0.3, 0.8)
    amp_anual: tuple[float, float] = (0.10, 0.30)
    promo_lambda: float = 1.5
    promo_boost: tuple[float, float] = (10.0, 40.0)
    intermitencia: float = 0.0                          # prob. de día en cero
    ruido: tuple[float, float] = (0.05, 0.15)           # sigma relativo (multiplicativo)
    event_boost: tuple[float, float] = (60.0, 220.0)

    @property
    def filas_history(self) -> int:
        return self.n_tiendas * self.n_productos * self.n_dias

    @property
    def filas_extra(self) -> int:
        """Filas de la hoja específica del dominio (una por serie en compras/inventario)."""
        if self.dominio in ("purchases", "inventory"):
            return self.n_tiendas * self.n_productos
        return 0


# ---------------------------------------------------------------------------
# El MANIFIESTO: perfiles distintos por tamaño, patrón y uso.
# ---------------------------------------------------------------------------
PERFILES: tuple[Perfil, ...] = (
    # ---- VENTAS -----------------------------------------------------------
    Perfil(
        "ventas_online_pequeno", "sales", 1, 1, 150, "2017-01-01", "online",
        "online", "Una serie con estacionalidad semanal/anual clara y tendencia leve.",
        granularity="day", horizon=7,
        nivel=(400, 600), tendencia=(0.15, 0.25), amp_semanal=(0.5, 0.6),
        amp_anual=(0.15, 0.2), promo_lambda=1.5, promo_boost=(25, 35), ruido=(0.08, 0.12),
    ),
    Perfil(
        "ventas_online_medio", "sales", 2, 3, 200, "2016-06-01", "online",
        "online", "6 series multi-tienda/producto con promociones frecuentes.",
        granularity="week", horizon=8,
        nivel=(150, 1200), tendencia=(-0.05, 0.35), amp_semanal=(0.3, 0.8),
        amp_anual=(0.1, 0.3), promo_lambda=2.2, promo_boost=(15, 45), intermitencia=0.04,
    ),
    Perfil(
        "ventas_frontera_bajo", "sales", 1, 1, 2000, "2012-07-05", "online",
        "frontera", "Serie larga de 2000 días = el MÁXIMO que aún se procesa en línea.",
        granularity="day", horizon=30,
        nivel=(500, 700), tendencia=(0.1, 0.3), amp_semanal=(0.4, 0.6),
        amp_anual=(0.15, 0.25), promo_lambda=1.8, promo_boost=(20, 40),
    ),
    Perfil(
        "ventas_frontera_alto", "sales", 1, 1, 2001, "2012-07-04", "batch",
        "frontera", "La misma serie + 1 fila (2001 días) = el MÍNIMO que cae en lote.",
        granularity="day", horizon=30,
        nivel=(500, 700), tendencia=(0.1, 0.3), amp_semanal=(0.4, 0.6),
        amp_anual=(0.15, 0.25), promo_lambda=1.8, promo_boost=(20, 40),
    ),
    Perfil(
        "ventas_lote_grande", "sales", 5, 10, 200, "2017-01-01", "batch",
        "lote", "Parque mediano (50 series) con granularidad mensual.",
        granularity="month", horizon=6,
        nivel=(100, 1500), tendencia=(-0.1, 0.4), amp_semanal=(0.3, 0.8),
        amp_anual=(0.1, 0.3), promo_lambda=2.0, promo_boost=(10, 50), intermitencia=0.05,
    ),
    Perfil(
        "ventas_lote_masivo", "sales", 10, 25, 200, "2017-01-01", "batch",
        "lote/estrés", "Parque grande (250 series, 50.000 filas) por debajo de 25 MB.",
        granularity="day", horizon=30,
        nivel=(50, 2000), tendencia=(-0.15, 0.45), amp_semanal=(0.2, 0.9),
        amp_anual=(0.05, 0.35), promo_lambda=2.5, promo_boost=(10, 60), intermitencia=0.06,
    ),
    Perfil(
        "ventas_retrain_rico", "sales", 5, 8, 730, "2015-01-01", "batch",
        "reentrenamiento", "40 series con 2 años de historia: permite un split temporal con sentido.",
        granularity="day", horizon=28,
        nivel=(120, 1400), tendencia=(-0.1, 0.5), amp_semanal=(0.3, 0.8),
        amp_anual=(0.15, 0.35), promo_lambda=2.2, promo_boost=(15, 55), intermitencia=0.04,
    ),
    Perfil(
        "ventas_retrain_escaso", "sales", 1, 1, 80, "2017-05-01", "online",
        "reentrenamiento", "Solo 80 días: caso HONESTO de 'no alcanza' para reentrenar.",
        granularity="day", horizon=7,
        nivel=(300, 500), tendencia=(0.0, 0.2), amp_semanal=(0.4, 0.6),
        amp_anual=(0.1, 0.15), promo_lambda=1.2, promo_boost=(20, 30),
    ),
    Perfil(
        "ventas_retrain_intermitente", "sales", 3, 5, 200, "2016-06-01", "batch",
        "reentrenamiento", "15 series con MUCHOS ceros (demanda intermitente): robustez.",
        granularity="day", horizon=14,
        nivel=(20, 120), tendencia=(-0.1, 0.2), amp_semanal=(0.2, 0.5),
        amp_anual=(0.05, 0.15), promo_lambda=0.6, promo_boost=(5, 20), intermitencia=0.55,
    ),
    # ---- COMPRAS ----------------------------------------------------------
    Perfil(
        "compras_pequeno", "purchases", 2, 3, 120, "2017-03-01", "online",
        "online", "6 series + replenishment_params válidos (en línea).",
        nivel=(200, 900), tendencia=(0.0, 0.3), amp_semanal=(0.3, 0.6),
        amp_anual=(0.1, 0.2), promo_lambda=1.8, promo_boost=(15, 40),
    ),
    Perfil(
        "compras_grande", "purchases", 5, 8, 100, "2017-01-01", "batch",
        "lote", "40 series + replenishment_params válidos (lote).",
        nivel=(100, 1300), tendencia=(-0.1, 0.4), amp_semanal=(0.3, 0.8),
        amp_anual=(0.1, 0.3), promo_lambda=2.2, promo_boost=(10, 50), intermitencia=0.05,
    ),
    # ---- INVENTARIO -------------------------------------------------------
    Perfil(
        "inventario_pequeno", "inventory", 2, 3, 120, "2017-03-01", "online",
        "online", "6 series + inventory_status válidos (en línea).",
        nivel=(200, 900), tendencia=(0.0, 0.3), amp_semanal=(0.3, 0.6),
        amp_anual=(0.1, 0.2), promo_lambda=1.8, promo_boost=(15, 40),
    ),
    Perfil(
        "inventario_grande", "inventory", 5, 8, 100, "2017-01-01", "batch",
        "lote", "40 series + inventory_status válidos (lote).",
        nivel=(100, 1300), tendencia=(-0.1, 0.4), amp_semanal=(0.3, 0.8),
        amp_anual=(0.1, 0.3), promo_lambda=2.2, promo_boost=(10, 50), intermitencia=0.05,
    ),
)


# ---------------------------------------------------------------------------
# Construcción de una serie de tiempo realista
# ---------------------------------------------------------------------------
def _rng(seed: int, *etiquetas: int) -> np.random.Generator:
    """Generador determinista derivado de la semilla global y unas etiquetas enteras."""
    return np.random.default_rng([seed, *etiquetas])


def _entre(rng: np.random.Generator, rango: tuple[float, float]) -> float:
    return float(rng.uniform(rango[0], rango[1]))


def _fechas(inicio: str, n_dias: int) -> list[date]:
    d0 = date.fromisoformat(inicio)
    return [d0 + timedelta(days=i) for i in range(n_dias)]


def _serie(perfil: Perfil, fechas: list[date], rng: np.random.Generator) -> dict[str, np.ndarray]:
    """Genera la demanda diaria de UNA serie (más promo/transactions/eventos).

    Modelo (multiplicativo en el nivel, aditivo en los picos)::

        demanda = nivel · tendencia(t) · semanal(dow) · anual(doy) · (1+ruido)
                  + promo_boost·on_promotion + event_boost·evento
    """
    n = len(fechas)
    t = np.arange(n, dtype=float)
    dow = np.array([f.weekday() for f in fechas])
    doy = np.array([f.timetuple().tm_yday for f in fechas], dtype=float)

    nivel = _entre(rng, perfil.nivel)
    tendencia = _entre(rng, perfil.tendencia)
    amp_sem = _entre(rng, perfil.amp_semanal)
    amp_anu = _entre(rng, perfil.amp_anual)
    ruido = _entre(rng, perfil.ruido)
    fase = float(rng.uniform(0, 2 * np.pi))
    pboost = _entre(rng, perfil.promo_boost)
    eboost = _entre(rng, perfil.event_boost)

    f_tendencia = 1.0 + tendencia * (t / max(1, n - 1))
    f_semanal = 1.0 + amp_sem * _DOW_FACTOR[dow]
    f_anual = 1.0 + amp_anu * np.sin(2 * np.pi * doy / 365.25 + fase)

    promo = rng.poisson(perfil.promo_lambda, n).astype(int)
    evento = np.array([(f.month, f.day) in _EVENTOS for f in fechas])
    factor_ruido = 1.0 + rng.normal(0.0, ruido, n)

    demanda = nivel * f_tendencia * f_semanal * f_anual * factor_ruido
    demanda = demanda + pboost * promo + eboost * evento
    demanda = np.maximum(0.0, demanda)

    if perfil.intermitencia > 0:
        cero = rng.random(n) < perfil.intermitencia
        demanda = np.where(cero, 0.0, demanda)

    # transactions correlacionada con la demanda (≥0); null en días sin venta.
    ratio = float(rng.uniform(1.1, 1.8))
    transacciones = np.maximum(0.0, demanda * ratio + rng.normal(0, 5, n))

    return {
        "demanda": demanda,
        "promo": promo,
        "transacciones": transacciones,
        "evento": evento,
    }


def _filas_history(perfil: Perfil, seed: int, perfil_idx: int) -> tuple[list[dict[str, Any]], dict[tuple[str, str], float]]:
    """Construye las filas del bloque ``history`` y la demanda media reciente por serie."""
    fechas = _fechas(perfil.fecha_inicio, perfil.n_dias)
    productos = FAMILIAS[: perfil.n_productos]
    filas: list[dict[str, Any]] = []
    media_reciente: dict[tuple[str, str], float] = {}

    for s in range(1, perfil.n_tiendas + 1):
        store_id = str(s)
        for p_idx, producto in enumerate(productos):
            rng = _rng(seed, perfil_idx, s, p_idx)
            serie = _serie(perfil, fechas, rng)
            demanda = serie["demanda"]
            promo = serie["promo"]
            trans = serie["transacciones"]
            evento = serie["evento"]
            for i, f in enumerate(fechas):
                u = round(float(demanda[i]), 1)
                fila: dict[str, Any] = {
                    "date": f,
                    "store_id": store_id,
                    "product_id": producto,
                    "units_sold": u,
                    "on_promotion": int(promo[i]),
                    # null en días sin venta de series intermitentes (degradación elegante).
                    "transactions": (None if (u == 0.0 and perfil.intermitencia > 0)
                                     else round(float(trans[i]), 0)),
                    # solo marcamos los días de evento; el resto queda vacío (opcional).
                    "event_active": (True if bool(evento[i]) else None),
                }
                filas.append(fila)
            ventana = demanda[-28:] if len(demanda) >= 28 else demanda
            media_reciente[(store_id, producto)] = float(np.mean(ventana)) if len(ventana) else 0.0

    return filas, media_reciente


def _filas_dominio(perfil: Perfil, media_reciente: dict[tuple[str, str], float], seed: int, perfil_idx: int) -> list[dict[str, Any]]:
    """Filas de la hoja específica del dominio (replenishment_params / inventory_status)."""
    if perfil.dominio not in ("purchases", "inventory"):
        return []
    rng = _rng(seed, perfil_idx, 9999)
    filas: list[dict[str, Any]] = []
    for (store_id, producto), media in media_reciente.items():
        current_stock = round(max(0.0, media * float(rng.uniform(2.0, 8.0))), 1)
        lead = int(rng.integers(1, 15))  # 1..14 (>0)
        if perfil.dominio == "purchases":
            filas.append({
                "store_id": store_id,
                "product_id": producto,
                "current_stock": current_stock,
                "lead_time_days": lead,
                "target_coverage_days": int(rng.integers(3, 31)),  # 3..30 (>0)
            })
        else:  # inventory: lead_time_days es opcional → ~30% vacío.
            filas.append({
                "store_id": store_id,
                "product_id": producto,
                "current_stock": current_stock,
                "lead_time_days": (None if rng.random() < 0.3 else lead),
            })
    return filas


# ---------------------------------------------------------------------------
# Escritura del .xlsx con las hojas/columnas canónicas del sistema
# ---------------------------------------------------------------------------
def _escribir(plantilla: PlantillaDominio, filas_por_hoja: dict[str, list[dict[str, Any]]], destino: Path) -> None:
    """Escribe un workbook con EXACTAMENTE las hojas/columnas del contrato (cabecera inglesa).

    Usa el modo normal de openpyxl (no ``write_only``): así las celdas vacías finales
    (p. ej. ``transactions``/``event_active`` ausentes) quedan en la dimensión de la
    hoja y el lector lee filas de ancho consistente, igual que con la plantilla real.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for hoja in plantilla.hojas:
        ws = wb.create_sheet(title=hoja.nombre)
        columnas = [c.nombre for c in hoja.columnas]
        ws.append(columnas)
        for fila in filas_por_hoja.get(hoja.nombre, []):
            ws.append([fila.get(c) for c in columnas])
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)


def _hoja_lista(plantilla: PlantillaDominio) -> HojaExcel:
    """Devuelve la hoja-lista específica del dominio (no la de history)."""
    for hoja in plantilla.hojas:
        if hoja.es_lista and hoja.nombre != "history":
            return hoja
    raise KeyError("sin hoja-lista de dominio")


def generar_archivo(perfil: Perfil, seed: int, perfil_idx: int, out_dir: Path) -> Path:
    """Genera el ``.xlsx`` de un perfil y lo guarda en ``out_dir``. Devuelve la ruta."""
    plantilla = plantilla_de(perfil.dominio)
    filas_history, media_reciente = _filas_history(perfil, seed, perfil_idx)
    filas_por_hoja: dict[str, list[dict[str, Any]]] = {"history": filas_history}

    if perfil.dominio == "sales":
        filas_por_hoja["parameters"] = [{
            "granularity": perfil.granularity,
            "horizon": perfil.horizon,
        }]
    else:
        hoja = _hoja_lista(plantilla)
        filas_por_hoja[hoja.nombre] = _filas_dominio(perfil, media_reciente, seed, perfil_idx)

    destino = out_dir / f"{perfil.nombre}.xlsx"
    _escribir(plantilla, filas_por_hoja, destino)
    return destino


# ---------------------------------------------------------------------------
# Archivos a propósito MAL formados (demostración del manejo de errores → 422)
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Invalido:
    nombre: str
    motivo: str


INVALIDOS: tuple[Invalido, ...] = (
    Invalido("mal_texto_en_numero", "Texto ('N/D') en la columna numérica units_sold."),
    Invalido("mal_falta_columna", "Falta la columna obligatoria units_sold."),
    Invalido("mal_obligatorio_vacio", "Celda vacía en el campo obligatorio store_id."),
)


def _base_history_invalida(seed: int) -> list[dict[str, Any]]:
    """Un history de ventas pequeño y por lo demás correcto, para luego corromperlo."""
    perfil = Perfil(
        "_base_invalida", "sales", 1, 1, 20, "2017-01-01", "online", "—", "—",
        granularity="day", horizon=7,
    )
    filas, _ = _filas_history(perfil, seed, 777)
    return filas


def generar_invalidos(seed: int, out_dir: Path) -> list[tuple[str, Path]]:
    """Genera los archivos mal formados en ``out_dir`` (carpeta separada). Devuelve (motivo, ruta)."""
    plantilla = plantilla_de("sales")
    cols_history = [c.nombre for c in plantilla.hojas[0].columnas]
    parameters = [{"granularity": "day", "horizon": 7}]
    resultados: list[tuple[str, Path]] = []

    for inv in INVALIDOS:
        filas = [dict(f) for f in _base_history_invalida(seed)]
        cols = list(cols_history)
        if inv.nombre == "mal_texto_en_numero":
            filas[3]["units_sold"] = "N/D"  # texto en columna numérica
        elif inv.nombre == "mal_falta_columna":
            cols = [c for c in cols if c != "units_sold"]  # falta columna obligatoria
        elif inv.nombre == "mal_obligatorio_vacio":
            filas[2]["store_id"] = None  # obligatorio vacío

        wb = Workbook()
        wb.remove(wb.active)
        ws_h = wb.create_sheet(title="history")
        ws_h.append(cols)
        for f in filas:
            ws_h.append([f.get(c) for c in cols])
        ws_p = wb.create_sheet(title="parameters")
        ws_p.append(["granularity", "horizon"])
        for f in parameters:
            ws_p.append([f["granularity"], f["horizon"]])

        destino = out_dir / f"{inv.nombre}.xlsx"
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)
        resultados.append((inv.motivo, destino))
    return resultados


# ---------------------------------------------------------------------------
# Manifiesto (CSV) con los metadatos de cada archivo
# ---------------------------------------------------------------------------
def escribir_manifiesto(out_dir: Path) -> Path:
    """Escribe ``MANIFIESTO.csv`` con el perfil y uso de cada archivo válido."""
    ruta = out_dir / "MANIFIESTO.csv"
    campos = [
        "archivo", "dominio", "tiendas", "productos", "dias", "filas_history",
        "filas_extra", "modo_esperado", "uso", "granularity", "horizon", "descripcion",
    ]
    with ruta.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=campos)
        w.writeheader()
        for p in PERFILES:
            w.writerow({
                "archivo": f"{p.nombre}.xlsx",
                "dominio": p.dominio,
                "tiendas": p.n_tiendas,
                "productos": p.n_productos,
                "dias": p.n_dias,
                "filas_history": p.filas_history,
                "filas_extra": p.filas_extra,
                "modo_esperado": p.modo_esperado,
                "uso": p.uso,
                "granularity": p.granularity or "",
                "horizon": p.horizon or "",
                "descripcion": p.descripcion,
            })
    return ruta


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Genera Excels sintéticos conformes al contrato SPC.")
    parser.add_argument("--out", default="data/synthetic", help="Carpeta de salida (default: data/synthetic, ya gitignored).")
    parser.add_argument("--seed", type=int, default=42, help="Semilla global (default: 42).")
    parser.add_argument("--sin-invalidos", action="store_true", help="No generar los archivos mal formados.")
    args = parser.parse_args(argv)

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Generando en {out_dir} con semilla {args.seed}\n")
    for idx, perfil in enumerate(PERFILES):
        ruta = generar_archivo(perfil, args.seed, idx, out_dir)
        mb = ruta.stat().st_size / (1024 * 1024)
        print(f"  [ok] {ruta.name:32s} {perfil.filas_history:>6d} filas  "
              f"{perfil.modo_esperado:6s}  {mb:5.2f} MB")

    manifiesto = escribir_manifiesto(out_dir)
    print(f"\n  [ok] {manifiesto.name} (manifiesto de {len(PERFILES)} archivos)")

    if not args.sin_invalidos:
        inv_dir = out_dir / "invalidos"
        print(f"\nGenerando archivos MAL formados en {inv_dir} (para demostrar el manejo de errores):")
        for motivo, ruta in generar_invalidos(args.seed, inv_dir):
            print(f"  [x] {ruta.name:28s} -> {motivo}")

    print("\nListo. Valida con: python scripts/validar_sinteticos.py --dir", out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
